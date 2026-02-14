"""
Handles generic input/output file writes
does NOT handle specific marc_21 marc files or excel io
(These are currently still within marc_21.py)
"""

from tkinter import W
import yaml
from pathlib import Path
import csv
import openpyxl  # type: ignore
import openpyxl.styles
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from datetime import date
import logging

from art_cats.settings import Default_settings

logger = logging.getLogger(__name__)

def get_base_filename(filepath: Path) -> str:
    base = f"{filepath.stem}.new{filepath.suffix}"
    print(f"{filepath.stem=}, {filepath.suffix=}")
    return base

def save_as_yaml(file: str, data) -> None:
    with open(file, mode="wt", encoding="utf-8") as f:
        yaml.dump(data, f, sort_keys=False)


# def open_yaml_file(file:str):
def open_yaml_file(file_path: Path):
    # with open(file, mode="rt", encoding="utf-8") as f:
    # file_path = settings.files.app_dir / Path(file)
    # file_path = Path(file)
    with open(file_path, mode="rt", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_plaintext_from_file(file_name: str) -> str:
    """Reads the plaintext content of the specified file, returning a default message on error. The plaintext could also encode markdown or html (as is the case here)."""
    path = Path(file_name)
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except IOError as e:
            return f"<h1>Error loading help content!</h1><p>Could not read file: {path}. Error: {e}</p>"
    else:
        return f"<h1>Help File Not Found</h1><p>Please create a file named '<b>{file_name}</b>' in the current directory.</p>"


def write_to_csv(file_name: Path, data: list[list[str]], headers: list[str]) -> None:
    # def write_to_csv(file_name: str, data: list[list[str]], headers: list[str]) -> None:
    # out_file = Path(settings.files.full_output_dir) / Path(file_name)
    # with open(out_file, "w", newline="", encoding="utf-8") as f:
    logger.info(f"Exporting records as csv to {file_name}")
    with open(file_name, "w", newline="", encoding="utf-8") as f:
        csvwriter = csv.writer(f)
        csvwriter.writerow(headers)
        csvwriter.writerows(data)


def get_csv_file_name_and_path(live_settings: Default_settings) -> Path:
    csv_file = (
        # live_settings.files.full_output_dir / f"{live_settings.files.out_file}.csv"
        live_settings.files.full_output_dir / live_settings.files.out_file
    )
    return csv_file

## +++++++++++ from marc_21.py


def extract_from_excel(excel_sheet, first_row_is_header:bool) -> tuple[list[str], list[list[str]]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    sheet = []
    headers = []
    # for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
    for i, excel_row in enumerate(excel_sheet.iter_rows(min_row=1, values_only=True)):
        if not excel_row[0] and not excel_row[1]:
            break  ## needed as openpyxl keeps spitting out empty rows at the end
        row = normalize_row(excel_row)
        # if i == 0 and settings.first_row_is_header:
        if i == 0 and first_row_is_header:
            headers = row
        else:
            sheet.append(row)
    return (headers, sheet)


def normalize_row(row: list) -> list:
    clean_row = []
    for col in row:
        if col:
            data = str(col).strip()
            data = trim_mistaken_decimals(data)
        else:
            data = ""
        clean_row.append(data)
    return clean_row


# def trim_mistaken_decimals(value: str | int) -> str:
#     if not isinstance(value, str):
#         value = str(value)
def trim_mistaken_decimals(value: str) -> str:
    if value.endswith(".0"):
        value = value[:-2]
    return value


def parse_file_into_rows(
    file_path: Path,
    first_row_is_header
) -> tuple[list[str], list[list[str]]]:
    is_excel_file = file_path.suffix.startswith(".xl")
    if is_excel_file:
        excel_file_name = str(file_path.resolve())
        worksheet = openpyxl.load_workbook(filename=excel_file_name).active
        headers, raw_rows = extract_from_excel(worksheet, first_row_is_header)
    else:
        headers, raw_rows = extract_from_csv(file_path, first_row_is_header)
    return (headers, raw_rows)


def extract_from_csv(file_address: Path, first_row_is_header) -> tuple[list[str], list[list[str]]]:
    sheet = []
    headers = []
    delimiter = "," if file_address.suffix == ".csv" else "\t"
    with open(file_address.resolve(), mode="r", encoding="utf-8") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        for i, row in enumerate(csv_reader):
            row = normalize_row(row)
            # if i == 0 and settings.first_row_is_header:
            if i == 0 and first_row_is_header:
                headers = row
            else:
                sheet.append(row)
    return (headers, sheet)


def write_CHU_file(
        chu_rows:list[list[str]],
        file_name: Path,
        ) -> None:
    """
    Write out CHU file, including formatting (for the craic)
    """
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore
    ws.title = "Recorded data"
    dark_blue = "24069B"
    lighter_dark_blue = "366092"
    light_cyan = "D2EEE7"

    # Merge cells for the header title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Alma holdings information update form"

    # Style for the header
    header_font = openpyxl.styles.Font(
        name="Arial", size=16, bold=True, color=dark_blue
    )  # Dark blue
    header_fill = openpyxl.styles.PatternFill(
        start_color=light_cyan, end_color=light_cyan, fill_type="solid"
    )  # Light cyan
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    cell = ws["A1"]
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    ws.row_dimensions[1].height = 45

    # Sub-row details
    # today = date.today()
    # today = today.strftime("%d %b %Y")
    today = date.today().strftime("%d %b %Y")
    initials = "PTW"
    email = "paul.wakelin@bodleian.ox.ac.uk"
    ws["A2"] = f"Date: {today}"
    ws["C2"] = "Initials:"
    ws["C2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["D2"] = initials
    ws["E2"] = "Contact e-mail:"
    ws["E2"].alignment = openpyxl.styles.Alignment(horizontal="right")
    ws["F2"] = email
    ws["G2"] = "(Always e-mail)"

    ## These are required by the CHU process
    named_ranges = {
        "WhenEmail": "G2",
        "ContactEmail": "F2",
    }
    for name, coord in named_ranges.items():
        ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(coord)}"
        named_range = DefinedName(name=name, attr_text=ref)
        wb.defined_names.add(named_range)

    # for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    for col in ["A", "B", "C", "D", "E", "F"]:
        cell = ws[f"{col}2"]
        cell.fill = header_fill
        cell.font = openpyxl.styles.Font(
            name="Arial", size=8, bold=False, color=lighter_dark_blue
        )
        # cell.alignment = Alignment(horizontal="left")
    ws["F1"].fill = header_fill
    ws["G2"].fill = openpyxl.styles.PatternFill()
    ws["G2"].font = openpyxl.styles.Font(
        name="Arial", size=7, bold=False, color=lighter_dark_blue
    )
    ws.row_dimensions[2].height = 10  # Height in points

    # Adjust column widths for better layout
    ws.column_dimensions["A"].width = 12  # Barcode
    ws.column_dimensions["B"].width = 10  # Library
    ws.column_dimensions["C"].width = 20  # Location
    ws.column_dimensions["D"].width = 12  # Item policy
    ws.column_dimensions["E"].width = 20  # Process
    ws.column_dimensions["F"].width = 30  # Shelfmark

    default_row_height = 13
    headers = ["Barcode", "Library", "Location", "Item Policy", "Process", "Shelfmark"]
    # Write headers to row 3 (assuming rows 1 and 2 are for the title and sub-header)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = openpyxl.styles.Font(name="Arial", bold=True, size=10)
        cell.alignment = openpyxl.styles.Alignment(horizontal="left")
    ws.row_dimensions[3].height = default_row_height  # Height in points

    for row_count, row in enumerate(chu_rows, 4):
        # row = [r[barcode_index], "", "", "", "Relocating to CSF", ""]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_count, column=col, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal="left")
            cell.font = openpyxl.styles.Font(name="Arial", bold=False, size=10)
        ws.row_dimensions[row_count].height = default_row_height  # Height in points
    print(file_name)
    wb.save(file_name)


def write_data_to_excel(
    data: list[list[str]],
    filename=Path("output_data.xlsx"),
    sheet_name: str = "Sheet1",
) -> None:
    """
    Writes a list of lists (including headers) to a new Excel file using openpyxl.

    The first inner list is treated as the header row.
    Handles Unicode text, including Chinese characters.

    Args:
        data: The structured data as list[list[str]].
        filename: The name of the Excel file to create/overwrite.
        sheet_name: The name of the worksheet to write to.
    """

    if not data or not data[0]:
        print("Error: Input data is empty or malformed. Cannot create file.")
        return

    # 1. Create a new Workbook and get the active worksheet
    try:
        workbook = openpyxl.Workbook()
        sheet: Worksheet = workbook.active # type: ignore
        sheet.title = sheet_name

        # 2. Iterate through the rows in the data structure
        for row_index, row_data in enumerate(data):
            # Rows in openpyxl are 1-indexed, not 0-indexed
            excel_row_num = row_index + 1

            # 3. Iterate through the cells in the current row
            for col_index, cell_value in enumerate(row_data):
                # Columns are also 1-indexed
                excel_col_num = col_index + 1

                # Write the value to the cell
                sheet.cell(row=excel_row_num, column=excel_col_num, value=cell_value)

        # 4. Save the Workbook
        workbook.save(filename)
        print(f"Successfully wrote data to '{filename}' on sheet '{sheet_name}'.")

    except Exception as e:
        print(f"An error occurred while writing the Excel file: {e}")