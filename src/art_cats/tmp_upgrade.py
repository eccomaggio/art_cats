#!/usr/bin/env python3
"""
insert_illustrated.py

Standalone script (no pandas) that accepts a single .csv, .xlsx, or .xlsm file path.
If the file has exactly 28 columns, inserts a new column at index 19 (0-based).
 - If the file has a header row (CSV: detected with csv.Sniffer; Excel: heuristic), the inserted
   header value is "Illustrated".
 - For data rows: the inserted cell is "True" ONLY IF the row already contains some data in any
   of its original columns; otherwise the inserted cell is left empty.

The updated file is saved alongside the original with ".updated" inserted before the file suffix:
  myfile.csv  -> myfile.updated.csv
  sheet.xlsx  -> sheet.updated.xlsx
  macro.xlsm  -> macro.updated.xlsm

Dependencies:
  - openpyxl

Usage:
  python insert_illustrated.py /path/to/file.csv
"""
from pathlib import Path
import argparse
import sys
import csv
import shutil
import re
from openpyxl import load_workbook

# # Excel support
# try:
#     from openpyxl import load_workbook
# except Exception as e:
#     print(
#         "ERROR: This script requires openpyxl. Install with: pip install openpyxl",
#         file=sys.stderr,
#     )
#     raise

CSV_EXTS = {".csv"}
XLS_EXTS = {".xlsx", ".xlsm"}

INSERT_INDEX_ZERO_BASE = 19  # user requested index 19 (0-based)


def updated_filename(path: Path) -> Path:
    """Insert '.updated' before the final suffix (e.g., data.csv -> data.updated.csv)."""
    return path.with_name(f"{path.stem}.updated{path.suffix}")


def detect_csv_dialect_and_header(path: Path, sample_size: int = 8192):
    """
    Uses csv.Sniffer to detect dialect and whether there is a header.
    Returns (dialect, has_header_bool).
    If sniffing fails, returns default csv.excel dialect and assumes no header.
    """
    sample = ""
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(sample_size)
    except Exception:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(2048)

    sniffer = csv.Sniffer()
    dialect = csv.excel
    has_header = False
    try:
        dialect = sniffer.sniff(sample)
    except Exception:
        dialect = csv.excel

    try:
        has_header = sniffer.has_header(sample)
    except Exception:
        has_header = False

    return dialect, has_header


def row_has_data(row):
    """
    Return True if any cell in the row contains a non-empty value after stripping whitespace.
    Treats None as empty.
    """
    for c in row:
        if c is None:
            continue
        if isinstance(c, str):
            if c.strip() != "":
                return True
        else:
            # Non-string (numbers, booleans) count as data
            return True
    return False


def process_csv(path: Path, outpath: Path):
    # detect dialect and header using a sample
    dialect, detected_header = detect_csv_dialect_and_header(path)
    rows = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, dialect)
        for r in reader:
            # Keep cell values as strings; None shouldn't appear from csv.reader
            rows.append([("" if c is None else str(c)) for c in r])

    if not rows:
        print(f"Input CSV is empty. Writing an updated copy to {outpath}")
        outpath.write_text("", encoding="utf-8")
        return

    # Determine number of columns based on maximum row length (conservative)
    max_cols = max(len(r) for r in rows)
    if max_cols != 28:
        print(
            f"No change: CSV has {max_cols} columns (expected 28). Writing a copy to: {outpath}"
        )
        shutil.copy2(path, outpath)
        return

    insert_at = INSERT_INDEX_ZERO_BASE
    has_header = detected_header

    new_rows = []
    for idx, original_row in enumerate(rows):
        # normalize to exactly max_cols columns
        row = list(original_row) + [""] * (max_cols - len(original_row))

        if has_header and idx == 0:
            insert_value = "Illustrated"
        else:
            # Only insert "True" if there is any existing data in the original row
            # Use the original row content (before insertion) to decide
            if row_has_data(row):
                insert_value = "True"
            else:
                insert_value = ""  # leave blank when row is empty
        new_row = row[:insert_at] + [insert_value] + row[insert_at:]
        new_rows.append(new_row)

    # Write using the detected dialect
    with outpath.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, dialect)
        for r in new_rows:
            writer.writerow(r)

    print(f"Updated CSV written to: {outpath} (inserted column at index {insert_at})")


def excel_first_row_has_letters(values):
    """
    Heuristic for Excel header detection: return True if any cell in first row contains ASCII letters.
    Accepts an iterable of cell values (which may be None).
    """
    for v in values:
        if v is None:
            continue
        s = str(v)
        if re.search(r"[A-Za-z]", s):
            return True
    return False


def process_excel(path: Path, outpath: Path):
    keep_vba = path.suffix.lower() == ".xlsm"
    try:
        wb = load_workbook(filename=str(path), data_only=False, keep_vba=keep_vba)
    except Exception as e:
        print(f"ERROR: could not open Excel file {path}: {e}", file=sys.stderr)
        raise

    ws = wb.worksheets[0]

    # Collect original row values (values_only) to inspect data presence before modifying the sheet
    rows_values = []
    actual_max = 0
    for row in ws.iter_rows(values_only=True):
        rows_values.append(list(row))
        actual_max = max(actual_max, len(row))

    max_cols = max(actual_max, ws.max_column)

    if max_cols != 28:
        print(
            f"No change: Excel sheet has {max_cols} columns (expected 28). Saving copy to: {outpath}"
        )
        wb.save(str(outpath))
        return

    insert_at_1_based = INSERT_INDEX_ZERO_BASE + 1
    # Insert the column (this shifts the sheet to the right)
    ws.insert_cols(idx=insert_at_1_based)

    # Header detection: check first original-row values
    first_row_vals = rows_values[0] if rows_values else [None] * max_cols
    has_header = excel_first_row_has_letters(first_row_vals)

    total_rows = ws.max_row
    # Fill the inserted column based on header and whether that row had data originally.
    # Note: rows_values corresponds to original rows before insertion; it may be shorter than total_rows if sheet had trailing empty rows.
    for row_idx in range(1, total_rows + 1):
        cell = ws.cell(row=row_idx, column=insert_at_1_based)
        if has_header and row_idx == 1:
            cell.value = "Illustrated"
        else:
            # Determine original row index in rows_values (0-based)
            orig_idx = row_idx - 1
            orig_row = rows_values[orig_idx] if orig_idx < len(rows_values) else []
            # Normalize orig_row length to max_cols for checking
            orig_row_norm = list(orig_row) + [None] * (max_cols - len(orig_row))
            if row_has_data(orig_row_norm):
                cell.value = "True"
            else:
                cell.value = None  # leave blank

    try:
        wb.save(str(outpath))
    except Exception as e:
        print(f"ERROR saving Excel file to {outpath}: {e}", file=sys.stderr)
        raise

    print(
        f"Updated Excel written to: {outpath} (inserted column at index {INSERT_INDEX_ZERO_BASE})"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Insert 'Illustrated' column at index 19 when file has 28 columns."
    )
    parser.add_argument(
        "file", metavar="FILE", help="Path to .csv, .xlsx, or .xlsm file"
    )
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"ERROR: File not found: {path}", file=sys.stderr)
        sys.exit(2)

    ext = path.suffix.lower()
    outpath = updated_filename(path)

    try:
        if ext in CSV_EXTS:
            process_csv(path, outpath)
        elif ext in XLS_EXTS:
            process_excel(path, outpath)
        else:
            print(
                "ERROR: Unsupported file type. Only .csv, .xlsx and .xlsm supported.",
                file=sys.stderr,
            )
            sys.exit(3)
    except Exception as exc:
        print(f"ERROR processing file: {exc}", file=sys.stderr)
        sys.exit(4)


if __name__ == "__main__":
    main()
