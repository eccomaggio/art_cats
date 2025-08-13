from openpyxl import load_workbook  # type: ignore
from dataclasses import dataclass, fields
from abc import ABC, abstractmethod
from typing import List, TypeAlias, Callable
from pprint import pprint
import sys
from datetime import datetime, timezone
import re
from pathlib import Path
import logging
# from pyisbn import Isbn
from pydantic import BaseModel, field_validator, ValidationError
from pydantic_extra_types.isbn import ISBN

logging.basicConfig(
    filename="output.log",
    filemode="w",
    encoding="utf-8",
    format="%(levelname)s:%(message)s",
    level=logging.DEBUG
    )


class Isbn_check(BaseModel):
    isbn: ISBN | str = ""
    @field_validator("isbn", mode="before")
    @classmethod
    def validate_isbn_or_empty(cls, input: str | ISBN):
        if input == "":
            return input
        return ISBN(input)


class Barcode_check(BaseModel):
    """
    barcode standard is probably 'code 39 modulo 43', but a simple regex is enough as we never include the final checksum character
    """
    barcode: str 
    @field_validator("barcode")
    @classmethod
    def validate_barcode_or_empty(cls, barcode: str) -> str:
        barcode_pattern = r'^[367][0-9]{8}$'
        if not re.match(barcode_pattern, barcode):
            msg = f"Barcode {barcode} must be a 9-digit string starting with 3, 6, or 7."
            logger.critical(msg)
            raise ValueError(f"Barcode {barcode} must be a 9-digit string starting with 3, 6, or 7.")
        return barcode


class Serialisable(ABC):
    @abstractmethod
    def to_string(self) -> str:
        pass

    @abstractmethod
    def to_mrc(self) -> str:
        pass

# @dataclass
# class String(Serialisable):
#     def to_string(self) -> str:
#         return str(self)
    
#     def to_mrc(self) -> str:
#         return str(self)


@dataclass
class Blank(Serialisable):
    def to_string(self) -> str:
        return "\\"

    def to_mrc(self) -> str:
        return " "

    def render(self, blank_char = "\\") -> str:
        return blank_char


class Punctuation(Serialisable):
    """ISBD punctuation used between subfields
    https://www.itsmarc.com/crs/mergedprojects/lcri/lcri/1_0c__lcri.htm
    """
    # def __init__(self, contents: str, code: str|int = ""):
    def __init__(self, contents: str):
        self.contents = contents

    def to_string(self) -> str:
        return self.contents

    def to_mrc(self) -> str:
        return self.to_string()

    def __str__(self):
        return f"<contents='{self.contents}'>"


class Subfield(Serialisable):
    """Subfield class for MARC21 subfields.
    -1 means the content is a single string (for variable control fields)
    if code is omitted, assumed to be -1 (i.e. no code)
    """
    # def __init__(self, code: str | int, data: str):
    def __init__(self,  data: str, code: str | int = -1):
        self.code: str = code if isinstance(code, str) else str(code)
        self.data: str = data

    def to_string(self) -> str:
        """Returns the subfield as a string."""
        delimiter = "$"
        if self.code == "-1":
            prefix = ""
        else:
            prefix = delimiter + str(self.code)
        return f"{prefix}{self.data}"

    def to_mrc(self) -> str:
        US = chr(31)
        if self.code == -1:
            prefix = ""
        else:
            prefix = f"{US}{self.code}"
        return f"{prefix}{self.data}"

    def __repl__(self):
        return f"<code='{self.code}', data='{self.data}'>"


atoms: TypeAlias = Subfield | Punctuation | Blank
# atoms: TypeAlias = String | Subfield | Punctuation | Blank
# data: TypeAlias = list[atoms] | atoms
worksheet_row: TypeAlias = list[str]


class Content(Serialisable):
    # def __init__(self, contents: data):
    def __init__(self, contents: atoms | list[atoms]):
        self.contents = []
        if not isinstance(contents, list):
          contents = [contents]
        for el in contents:
            if type(el) not in [Subfield, Punctuation, Blank]:
                raise TypeError("Only subfields or punctuation allowed")
            else:
                self.contents.append(el)

    def can_accept_period(self) -> bool:
        return bool(self.to_string().rstrip()[-1] not in "?!.")

    def to_string(self) -> str:
        return "".join([el.to_string() for el in self.contents])

    def to_mrc(self) -> str:
        return "".join([el.to_mrc() for el in self.contents])

    # def add(self, content: data):
    def add(self, content: list[atoms] | atoms):
        if isinstance(content, list):
            self.contents.extend(content)
        else:
            self.contents.append(content)

    def __getitem__(self, index:int):
        return self.contents[index]

    def __str__(self):
        result = ", ".join([str(el) for el in self.contents])
        return f"[{result}]"


class Field(Serialisable):
    """Variable Control/Data Field class for MARC21 variable control fields.
    -1 or Blank = indicator placeholder (various realizations, e.g. space or backslash)
    -2 = no indicator"""
    def __init__(self, tag: int, i1: int | Blank, i2: int | Blank, contents: Content | atoms | list[atoms], ordering: int = 1):
        self.tag = tag
        self.i1: int | Blank  = Blank() if i1 == -1 else i1
        self.i2: int | Blank = Blank() if i2 == -1 else i2
        self.contents = contents if isinstance(contents, Content) else Content(contents)
        self.ordering: int = ordering ## sort is used to order the fields when there are more than one of the same tag

    def expand_indicators(self, indicator: int | Blank, blank_char="\\") -> str:
        if isinstance(indicator, Blank):
            expansion = indicator.render(blank_char)
        elif indicator == -2:
            expansion = ""
        else:
            expansion = str(indicator)
        return expansion

    def can_accept_period(self) -> bool:
        return self.contents.can_accept_period()


    def to_string(self) -> str:
        """Returns the variable control field as a string."""
        tag = "LDR" if self.tag == 0 else str(self.tag).zfill(3)
        i1 = self.expand_indicators(self.i1)
        i2 = self.expand_indicators(self.i2)
        contents = self.contents.to_string()
        return f"={tag}  {i1}{i2}{contents}"

    def to_mrc(self) -> str:
        RS = chr(30)
        i1 = self.expand_indicators(self.i1, " ")
        i2 = self.expand_indicators(self.i2, " ")
        contents = self.contents.to_mrc()
        return f"{i1}{i2}{contents}{RS}"

    def __repr__(self):
        return f"={self.tag}@{self.ordering} ({self.i1})({self.i2}) {self.contents.to_string()}"

# fields: TypeAlias = list[Field] | Field
marc_record: TypeAlias = list[Field]

@dataclass
class Title:
    original: str
    transliteration: str


@dataclass
class Record:
    sublib: str
    langs: List[str]
    isbn: str
    title: Title
    subtitle: Title
    parallel_title: Title
    parallel_subtitle: Title
    country: str
    place: str
    publisher: str
    pub_year: str
    copyright: str
    extent: str
    size: int
    series_title: str
    series_enum: str
    notes: str
    sales_code: str
    sale_dates: List[str]
    hol_notes: str
    donation: str
    barcode: str

    pub_year_is_approx: bool
    extent_is_approx: bool
    timestamp: datetime
    
    sequence_number: int
    links: List[Field | None]


@dataclass
class Result:
    """
    holds the successfully acquired data OR
    the MARC field tag (as an int) where the problem occurred and an error message
    """
    is_ok: list[Field] | Field | None    ## MARC field or list of same
    is_err:tuple[int, str] | None   ## [field number, error message]


class MissingFieldError(Exception):
    pass


def norm_langs(raw: str) -> list[str]:
    language_codes = {
        "english": "eng",
        "chinese": "chi",
        "german": "ger",
        "italian": "ita",
        "spanish": "spa",
        "french": "fre",
        "swedish": "swe",
        "danish": "dan",
        "norwegian": "nor",
        "dutch": "dut",
    }
    list_of_languages = []
    languages: List[str] = raw.replace(" ", "").lower().split("/")
    for language in languages:
        try:
            list_of_languages.append(language_codes[language])
        except KeyError as e:
            logger.warning(f"Warning: {e} is not a recognised language; it has been passed on unchanged.")
            list_of_languages.append(language)
    return list_of_languages


def norm_geographical_name(name: str) -> str:
    return re.sub(r"[\s\-']", "", name).lower()


def norm_country(country_raw: str) -> str:
    country_codes = {
        "algeria": "ae",
        "angola": "ao",
        "benin": "dm",
        "botswana": "bs",
        "burkinafaso": "uv",
        "burundi": "bd",
        "cameroon": "cm",
        "centralafricanrepublic": "cx",
        "chad": "cd",
        "congo": "cf",
        "democraticrepublicofcongo": "cg",
        "côtedivoire": "iv",
        "cotedivoire": "iv",
        "djibouti": "ft",
        "egypt": "ua",
        "equatorialguinea": "eg",
        "eritrea": "ea",
        "ethiopia": "et",
        "gabon": "go",
        "gambia": "gm",
        "ghana": "gh",
        "guinea": "gv",
        "guineabissau": "pg",
        "kenya": "ke",
        "lesotho": "lo",
        "liberia": "lb",
        "libya": "ly",
        "madagascar": "mg",
        "malawi": "mw",
        "mali": "ml",
        "mauritania": "mu",
        "morocco": "mr",
        "mozambique": "mz",
        "namibia": "sx",
        "niger": "ng",
        "nigeria": "nr",
        "rwanda": "rw",
        "saotomeandprincipe": "sf",
        "senegal": "sg",
        "sierraleone": "sl",
        "somalia": "so",
        "southafrica": "sa",
        "southsudan": "sd",
        "spanishnorthafrica": "sh",
        "sudan": "sj",
        "swaziland": "sq",
        "tanzania": "tz",
        "togo": "tg",
        "tunisia": "ti",
        "uganda": "ug",
        "westernsahara": "ss",
        "zambia": "za",
        "zimbabwe": "rh",
        "afghanistan": "af",
        "armenia": "ai",
        "republicofarmenia": "ar",
        "azerbaijan": "aj",
        "bahrain": "ba",
        "bangladesh": "bg",
        "bhutan": "bt",
        "brunei": "bx",
        "burma": "br",
        "cambodia": "cb",
        "china": "cc",
        "cyprus": "cy",
        "easttimor": "em",
        "gazastrip": "gz",
        "georgia": "gs",
        "georgianrepublic": "gs",
        "republicofgeorgia": "gs",
        "india": "ii",
        "indonesia": "io",
        "iran": "ir",
        "iraq": "iq",
        "israel": "is",
        "japan": "ja",
        "jordan": "jo",
        "kazakhstan": "kz",
        "northkorea": "kn",
        "korea": "ko",
        "southkorea": "ko",
        "kuwait": "ku",
        "kyrgyzstan": "kg",
        "laos": "ls",
        "lebanon": "le",
        "malaysia": "my",
        "mongolia": "mp",
        "nepal": "np",
        "oman": "mk",
        "pakistan": "pk",
        "papuanewguinea": "pp",
        "paracelislands": "pf",
        "philippines": "ph",
        "qatar": "qa",
        "saudiarabia": "su",
        "singapore": "si",
        "spratlyisland": "xp",
        "srilanka": "ce",
        "syria": "sy",
        "tajikistan": "ta",
        "thailand": "th",
        "turkey": "tu",
        "turkmenistan": "tk",
        "unitedarabemirates": "ts",
        "uae": "ts",
        "uzbekistan": "uz",
        "vietnam": "vm",
        "westbankofthejordanriver": "wj",
        "westbank": "wj",
        "yemen": "ye",
        "bermudaislands": "bm",
        "bermuda": "bm",
        "bouvetisland": "bv",
        "caboverde": "cv",
        "faroeislands": "fa",
        "faroes": "fa",
        "falklandislands": "fk",
        "falklands": "fk",
        "sainthelena": "xj",
        "southgeorgiaandthesouthsandwichislands": "xs",
        "southgeorgia": "xs",
        "southsandwichislands": "xs",
        "belize": "bh",
        "costarica": "cr",
        "elsalvador": "es",
        "guatemala": "gt",
        "honduras": "ho",
        "nicaragua": "nq",
        "panama": "pn",
        "albania": "aa",
        "andorra": "an",
        "austria": "au",
        "belarus": "bw",
        "belgium": "be",
        "bosniaandherzegovina": "bn",
        "bosnia": "bn",
        "bosniaherzegovina": "bn",
        "herzegovina": "bn",
        "bulgaria": "bu",
        "croatia": "ci",
        "czechrepublic": "xr",
        "czechia": "xr",
        "denmark": "dk",
        "estonia": "er",
        "finland": "fi",
        "france": "fr",
        "germany": "gw",
        "gibraltar": "gi",
        "greece": "gr",
        "guernsey": "gg",
        "hungary": "hu",
        "iceland": "ic",
        "ireland": "ie",
        "isleofman": "im",
        "italy": "it",
        "jersey": "je",
        "kosovo": "kv",
        "latvia": "lv",
        "liechtenstein": "lh",
        "lithuania": "li",
        "luxembourg": "lu",
        "macedonia": "xn",
        "malta": "mm",
        "montenegro": "mo",
        "moldova": "mv",
        "monaco": "mc",
        "netherlands": "ne",
        "norway": "no",
        "poland": "pl",
        "portugal": "po",
        "serbia": "rb",
        "romania": "rm",
        "russia": "ru",
        "russianfederation": "ru",
        "sanmarino": "sm",
        "slovakia": "xo",
        "slovenia": "xv",
        "spain": "sp",
        "sweden": "sw",
        "switzerland": "sz",
        "ukraine": "un",
        "vaticancity": "vc",
        "serbiaandmontenegro": "yu",
        "britishindianoceanterritory": "bi",
        "christmasisland": "xa",
        "cocosislands": "xb",
        "keelingislands": "xb",
        "comoros": "cq",
        "heardandmcdonaldislands": "hm",
        "maldives": "xc",
        "mauritius": "mf",
        "mayotte": "ot",
        "réunion": "re",
        "reunion": "re",
        "seychelles": "se",
        "americansamoa": "as",
        "cookislands": "cw",
        "fiji": "fj",
        "frenchpolynesia": "fp",
        "guam": "gu",
        "johnstonatoll": "ji",
        "kiribati": "gb",
        "marshallislands": "xe",
        "micronesia": "fm",
        "federatedstatesofmicronesia": "fm",
        "midwayislands": "xf",
        "nauru": "nu",
        "newcaledonia": "nl",
        "niue": "xh",
        "northernmarianaislands": "nw",
        "palau": "pw",
        "pitcairnisland": "pc",
        "samoa": "ws",
        "solomonislands": "bp",
        "tokelau": "tl",
        "tonga": "to",
        "tuvalu": "tv",
        "vanuatu": "nn",
        "wakeisland": "wk",
        "wallisandfutuna": "wf",
        "wallis": "wf",
        "futuna": "wf",
        "argentina": "ag",
        "bolivia": "bo",
        "brazil": "bl",
        "chile": "cl",
        "colombia": "ck",
        "ecuador": "ec",
        "frenchguiana": "fg",
        "guyana": "gy",
        "paraguay": "py",
        "peru": "pe",
        "surinam": "sr",
        "uruguay": "uy",
        "venezuela": "ve",
        "anguilla": "am",
        "antiguaandbarbuda": "aq",
        "antigua": "aq",
        "barbuda": "aq",
        "aruba": "aw",
        "bahamas": "bf",
        "barbados": "bb",
        "britishvirginislands": "vb",
        "caribbeannetherlands": "ca",
        "caymanislands": "cj",
        "cuba": "cu",
        "curaçao": "co",
        "curacao": "co",
        "dominica": "dq",
        "dominicanrepublic": "dr",
        "grenada": "gd",
        "guadeloupe": "gp",
        "haiti": "ht",
        "jamaica": "jm",
        "martinique": "mq",
        "montserrat": "mj",
        "puertorico": "pr",
        "saintbarthélemy": "sc",
        "saintbarthelemy": "sc",
        "saintkittsnevis": "xd",
        "saintkitts": "xd",
        "nevis": "xd",
        "saintlucia": "xk",
        "saintmartin": "st",
        "saintvincentandthegrenadines": "xm",
        "saintvincent": "xm",
        "thegrenadines": "xm",
        "grenadines": "xm",
        "sintmaarten": "sn",
        "trinidadandtobago": "tr",
        "trinidad": "tr",
        "tobago": "tr",
        "turksandcaicosislands": "tc",
        "virginislandsoftheunitedstates": "vi",
        "antarctica": "ay",
        "noplace": "xx",
        "unknown": "xx",
        "undetermined": "xx",
        "variousplaces": "vp",
        "various": "vp",
    }
    # normed_country = country.replace(" ", "").lower()
    # country = re.sub(r"[\s\-']", "", country_raw).lower()
    country = norm_geographical_name(country_raw)
    try:
        result = country_codes[country]
    except KeyError as e:
        if len({e}) < 4:
            logger.info(f"Advisory: assuming country name ({e}) has already been processed.")
        else:
            logger.warning(f"Warning: {e} is not a recognised country name; it has been passed on unchanged.")
        result = country_raw
    return result


def norm_place(place_raw: str) -> str:
    long_country_codes = {
        "england": "enk",
        "northernireland": "nik",
        "scotland": "stk",
        "wales": "wlk",

        "alberta": "abc",
        "britishcolumbia": "bcc",
        "bc": "bcc",
        "manitoba": "mbc",
        "newbrunswick": "nkc",
        "newfoundland": "nfc",
        "labrador": "nfc",
        "newfoundlandandlabrador": "nfc",
        "northwestterritories": "ntc",
        "novascotia": "nsc",
        "nunavut": "nuc",
        "ontario": "onc",
        "princeedwardisland": "pic",
        "québecprovince": "quc",
        "quebéc": "quc",
        "quebecprovince": "quc",
        "quebec": "quc",
        "saskatchewan": "snc",
        "yukonterritory": "ykc",
        "yukon": "ykc",

        "alabama": "alu",
        "alaska": "aku",
        "arizona": "azu",
        "arkansas": "aru",
        "california": "cau",
        "colorado": "cou",
        "connecticut": "ctu",
        "delaware": "deu",
        "districtofcolumbia": "dcu",
        "columbia": "dcu",
        "florida": "flu",
        "georgia": "gau",
        "hawaii": "hiu",
        "idaho": "idu",
        "illinois": "ilu",
        "indiana": "inu",
        "iowa": "iau",
        "kansas": "ksu",
        "kentucky": "kyu",
        "louisiana": "lau",
        "maine": "meu",
        "maryland": "mdu",
        "massachusetts": "mau",
        "michigan": "miu",
        "minnesota": "mnu",
        "mississippi": "msu",
        "missouri": "mou",
        "montana": "mtu",
        "nebraska": "nbu",
        "nevada": "nvu",
        "newhampshire": "nhu",
        "newjersey": "nju",
        "newmexico": "nmu",
        "newyork": "nyu",
        "newyorkstate": "nyu",
        "northcarolina": "ncu",
        "northdakota": "ndu",
        "ohio": "ohu",
        "oklahoma": "oku",
        "oregon": "oru",
        "pennsylvania": "pau",
        "rhodeisland": "riu",
        "southcarolina": "scu",
        "southdakota": "sdu",
        "tennessee": "tnu",
        "texas": "txu",
        "utah": "utu",
        "vermont": "vtu",
        "virginia": "vau",
        "washington": "wau",
        "washingtonstate": "wau",
        "westvirginia": "wvu",
        "wisconsin": "wiu",
        "wyoming": "wyu",

        "australiancapitalterritory": "aca",
        "queensland": "qea",
        "tasmania": "tma",
        "victoria": "vra",
        "westernaustralia": "wea",
        "newsouthwales": "xna",
        "northernterritory": "xoa",
        "southaustralia": "xra",
    }
    # place = place_raw.replace(" ", "").lower()
    place = norm_geographical_name(place_raw)
    try:
        result = long_country_codes[place]
    except KeyError as e:
        if len({e}) == 3:
            logger.info(f"Advisory: assuming place name ({e}) has already been processed.")
        else:
            logger.warning(f"Warning: {e} is not a recognised place name; it has been passed on unchanged.")
        result = place_raw
    return result


def get_long_country_code(country: str, place: str) -> str:
    ## USA & UK return a detailed 3-digit code based on local region
    return place.strip().lower() if len(place) == 3 else country


def check_mandatory_fields_exist(record: Record) -> bool:
    mandatory = [
        "sublib",
        "langs",
        "title",
        "country",
        "place",
        "publisher",
        "pub_year",
        "extent",
        "size",
        "sale_dates",
        "barcode",
    ]
    is_valid = False
    for i, field in enumerate(fields(record)):
        name = field.name
        is_valid = True
        if name in mandatory and not getattr(record, name):
            logger.warning(f"Record no. {i} is missing the mandatory field '{name}'.")
            is_valid = False
    return is_valid


def norm_dates(raw: str) -> list[str]:
    result = [date.strip() for date in raw.split(",")]
    return result


def norm_size(raw: str) -> int:
    raw = strip_unwanted(r"cm",raw)
    return int(raw)


def norm_pages(pages_raw: str) -> str:
    pages = strip_unwanted(r"pages|\[|\]", pages_raw)
    if "approximately" in pages:
        pages = re.sub(r"\s?approximately\s?", "", pages)
        pages = pages + "?"
    return pages


def norm_year(year_raw: str) -> str:
    year = strip_unwanted(r"[\[\]]", year_raw)
    return year


# TODO: write sensible validation
def norm_isbn(raw_isbn: str) -> str:
    isbn = raw_isbn
    return isbn


# TODO: write sensible validation
def norm_barcode(raw_barcode: str) -> str:
    barcode = raw_barcode
    return barcode


def strip_unwanted(pattern: str, raw: str) -> str:
  clean = re.sub(pattern, "", raw)
  return clean


def check_for_approx(raw_string: str) -> tuple[str, bool]:
    clean = str(raw_string).strip()
    if clean[-1] == "?":
        is_approx = True
        clean = clean[:-1].rstrip()
    else:
        is_approx = False
    clean = trim_mistaken_decimals(clean)
    return (clean, is_approx)


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string


# TODO: what should this return if nothing? None or []?
def fill_with_blanks(string: str, limit: int = 3) -> list[Blank]:
    return [Blank() for n in range(limit - len(string))]


def create_date_list(dates_raw: str) -> list[str]:
    dates_raw = re.sub(r"\s|\.0", "", dates_raw)
    dates = dates_raw.split(",")
    return dates


# def extract_from_excel(excel_sheet) -> list[list[str]]:
def extract_from_excel(excel_sheet) -> list[worksheet_row]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    sheet = []
    for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
        row = []
        if not excel_row[0]:
            break
        for col in excel_row:
            if col:
                data = str(col).strip()
                data = trim_mistaken_decimals(data)
            else:
                data = ""
            row.append(data)
        sheet.append(row)
    return sheet


# def parse_spreadsheet(sheet: list[list[str]]) -> list[Record]:
def parse_rows_into_records(sheet: list[worksheet_row]) -> list[Record]:
    current_time = datetime.now()
    records = []
    for row in sheet:
        record = parse_row(row, current_time)
        # validate(record)
        records.append(record)
        # pprint(record: Record) -> Result
    return records


def parse_row(row: list[str], current_time: datetime) -> Record:
    cols = iter(row)
    sublibrary = next(cols)
    langs = norm_langs(next(cols))
    isbn = norm_isbn(next(cols))
    # isbn = Isbn_check(isbn=next(cols))
    title = Title(next(cols), next(cols))
    subtitle = Title(next(cols), next(cols))
    parallel_title = Title(next(cols), next(cols))
    parallel_subtitle = Title(next(cols), next(cols))
    country = norm_country(next(cols))
    place = next(cols)
    publisher = next(cols)
    pub_date, pub_date_is_approx = check_for_approx(norm_year(next(cols)))
    copyright_ = next(cols).replace("©","").strip()
    extent, extent_is_approx = check_for_approx(norm_pages(next(cols)))
    size = norm_size(next(cols))
    series_title = next(cols)
    series_enum = next(cols)
    note = next(cols)
    sale_code = next(cols)
    date_of_sale = create_date_list(next(cols))
    hol_notes = next(cols)
    donation = next(cols)
    barcode = norm_barcode(next(cols))
    # barcode = Barcode_check(barcode=next(cols))

    record = Record(
        sublibrary,
        langs,
        isbn,
        title,
        subtitle,
        parallel_title,
        parallel_subtitle,
        country,
        place,
        publisher,
        pub_date,
        copyright_,
        extent,
        size,
        series_title,
        series_enum,
        note,
        sale_code,
        date_of_sale,
        hol_notes,
        donation,
        barcode,

        pub_date_is_approx,
        extent_is_approx,
        current_time,

        sequence_number = 1,
        links = [],
    )
    check_mandatory_fields_exist(record)
    return record



def build_leader(record: Record) -> Result:
    """leader (0 is only for sorting purposes; should read 'LDR')"""
    tag = 0
    i1, i2 = variable_control_field()
    # content = "00000nam a22000003i 4500"
    record_len_00 = "00000"  # placeholder for record length
    record_status_05 = "n"  # "n" for new record
    record_type_06 = "a"  # "a" for language material
    biblio_level_07 = "m"  # "m" for monograph
    type_of_ctrl_08 = " "  # no type of control
    char_coding_09 = "a"  # Unicode
    indicator_count_10 = "2"  # no indicators
    subfield_count_11 = "2"  # no subfields
    base_address_12 = "00000"  # placeholder for base address of data
    encoding_level_17 = "3" # "3" for abbrieviated level
    cat_conventions_18 = "i"  # ISBD punctuation included
    multipart_indic_19 = " "  # no multipart
    field_len_20 = "4"  # placeholder for length of field
    start_character_len_21 = "5"  # placeholder for length of starting character
    implementation_len_22 = "0"  # placeholder for implementation length
    undefined_23 = "0"  # placeholder for undefined length
    data: str = record_len_00 + record_status_05 + record_type_06 + biblio_level_07 + type_of_ctrl_08 + char_coding_09 + indicator_count_10 + subfield_count_11 + base_address_12 + encoding_level_17 + cat_conventions_18 + multipart_indic_19 + field_len_20 + start_character_len_21 + implementation_len_22 + undefined_23
    success = Field(tag, i1, i2, [Subfield(data)])
    return Result(success, None)


def build_005(record: Record) -> Result:
    """date & time of transaction
    "The date requires 8 numeric characters in the pattern yyyymmdd. The time requires 8 numeric characters in the pattern hhmmss.f, expressed in terms of the 24-hour (00-23) clock."
    """
    tag = 5
    i1, i2 = variable_control_field()
    standard_time = record.timestamp.now(timezone.utc)
    ## NB: python produces this format: YYYY-MM-DD HH:MM:SS.ffffff, e.g. 2020-09-30 12:37:55.713351
    timestamp = str(standard_time).translate(str.maketrans("", "", " -:"))[:16]
    result = Result(Field(tag, i1, i2, [Subfield(timestamp)]), None)
    return result


def build_008(record: Record) -> Result:
    """pub year & main language?"""
    tag = 8
    i1, i2 = variable_control_field()
    t = record.timestamp
    date_entered_on_file = Subfield(str(t.year)[2:] + str(t.month).zfill(2) + str(t.day).zfill(2))
    pub_status: Subfield = Subfield("s")
    date_1: Subfield = Subfield(record.pub_year)
    date_2: Subfield = Subfield(4 * "|")
    place_of_pub: list[atoms] = [Subfield(record.country), *fill_with_blanks(record.country)]
    books_configuration: list[atoms] = [Subfield(14*"|"), Blank(), Subfield(2*"|")]
    lang: list[atoms] = [Subfield(record.langs[0]), *fill_with_blanks(record.langs[0])]
    modified_and_cataloging: Subfield = Subfield(2*"|")
    content = Content([date_entered_on_file,  pub_status,  date_1,  date_2,  *place_of_pub,  *books_configuration,  *lang,  modified_and_cataloging])
    result = Result(Field(tag, i1, i2, content), None)
    return result

def build_033(record: Record) -> Result:
    """sales dates"""
    tag = 33
    i1 = 0 if len(record.sale_dates) == 1 else 1
    i2 = -1
    result = Result(Field(tag, i1, i2, [Subfield("a", date) for date in record.sale_dates]), None)
    return result


def build_040(record: Record) -> Result:
    """Cataloguing source (Oxford)"""
    tag = 40
    i1, i2 = -1, -1
    content = Content([Subfield("a", "UkOxU"), Subfield("b", "eng"), Subfield("e", "rda"), Subfield("c", "UkOxU")])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_245(record: Record) -> Result:
    """Title
    Field 245 ends with a period, even when another mark of punctuation is present, unless the last word in the field is an abbreviation, initial/letter, or data that ends with final punctuation.
    """
    tag = 245
    i1 = 0
    nonfiling = 0
    error = None
    has_chinese_title = bool(record.title.transliteration)
    if has_chinese_title:
        title, subtitle = record.title.transliteration, record.subtitle.transliteration
        linkage = deal_with_chinese_titles(record, record.title.original, record.subtitle.original, i1, nonfiling, tag)
    else:
        title, subtitle = record.title.original, record.subtitle.original
        nonfiling, title = check_for_nonfiling(title)
        linkage = None
    i2 = nonfiling
    if title:
        content = Content([])
        if linkage:
            content.add(linkage)
        content.add(Subfield("a", title))
        if subtitle:
           content.add([Punctuation(" :"), Subfield("b", subtitle)])
        if content.can_accept_period():
            content.add(Punctuation("."))

        result = Result(Field(tag, i1, i2, content), error)
    else:
        result = Result(None, (tag, ""))
    return result


def deal_with_chinese_titles(record: Record, title_original: str, subtitle_original: str | None, i1: int, i2: int, tag: int) -> Subfield:
    chinese_title = Content(Subfield("a", title_original))
    if subtitle_original:
        chinese_title.add([Punctuation(" :"), Subfield("b", subtitle_original)])
    if tag == 245 and chinese_title.can_accept_period():
        chinese_title.add(Punctuation("."))
    sequence_number = seq_num(record.sequence_number)
    linkage = Subfield(f"880-{sequence_number}", 6)
    # print(chinese_title)
    # print(chinese_title.to_string())
    build_880(record, chinese_title, i1, i2, tag, sequence_number)
    return linkage


def build_264(record: Record) -> Result:
    """publisher & copyright"""
    tag = 264
    i1 = -1
    i2 = 1  ## "Publication: Field contains a statement relating to the publication, release, or issuing of a resource."
    # i2 = 0  ## "Production: Field contains a statement relating to the production of a resource."
    error = None
    place = Subfield("a", record.place)
    publisher = Subfield("b", record.publisher)
    pub_year = Subfield("c", f"[{record.pub_year}?]" if record.pub_year_is_approx else record.pub_year)
    content = Field(tag, i1, i2, [place, Punctuation(" :"), publisher, Punctuation(","), pub_year], 1)
    if record.copyright:
        _copyright = Field(tag, i1, 4, [Subfield("c",f"\u00a9 {record.copyright}")], 2)
        result  = Result([content, _copyright], error)
    else:
        result = Result(content, error)
    return result


def build_300(record: Record) -> Result:
    """physical description"""
    tag = 300
    i1, i2 = -1, -1 ## "undefined"
    pages = Subfield("a", f"approximately {record.extent} pages" if record.extent_is_approx else f"{record.extent} pages")
    size = Subfield("c", f"{record.size} cm")
    content = Content([pages, Punctuation(" ;"), size])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_336(record: Record) -> Result:
    """content type (boilerplate)"""
    tag = 336
    content = Content([Subfield("a", "text"), Subfield("rdacontent", 2)])
    result = Result(Field(tag, -1, -1,content), None)
    return result


def build_337(record: Record) -> Result:
    """media type (boilerplate)"""
    tag = 337
    i1, i2 = -1, -1
    content = Content([Subfield("a", "unmediated"), Subfield("rdamedia", 2)])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_338(record: Record) -> Result:
    """carrier type (boilerplate)"""
    tag = 338
    i1, i2 = -1, -1
    content =  Content([Subfield("a", "volume"), Subfield("rdacarrier", 2)])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_876(record: Record) -> Result:
    """
    notes / donations / barcode
    mandatory because of barcode
    """
    tag = 876
    i1, i2 = -1, -1
    content = Content([Subfield("p", record.barcode)])
    if (donation := record.donation):
        content.add(Subfield("z", donation))
    if (notes := record.hol_notes):
        content.add(Subfield("z", notes))
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_904(record: Record) -> Result:
    """authority (boilerplate)"""
    tag = 904
    i1, i2 = -1, -1
    content =  Content(Subfield("a", "Oxford Local Record"))
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_020(record: Record) -> Result:  ##optional
    """isbn (if exists)"""
    tag = 20
    i1, i2 = -1, -1
    if record.isbn:
        result = Result(Field(tag, i1, i2, Content(Subfield("a", record.isbn))), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_024(record: Record) -> Result:  ##optional
    """sales code (if exists)"""
    tag = 24
    i1 = 8
    i2 = -1
    if record.sales_code:
        result = Result(Field(tag, i1, i2, Content(Subfield("a", record.sales_code))), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_041(record: Record) -> Result:  ##optional
    """language codes if not monolingual"""
    tag = 41
    # i1 = -1  ## "No information...as to whether the item is or includes a translation."
    i1 = 0  ## "Item not a translation/does not include a translation."
    i2 = -1  ## "(followed by) MARC language code"
    is_multi_lingual = len(record.langs) > 1
    if is_multi_lingual:
        ## OPTION 1
        # if record.title.transliteration:
        #     main_lang = record.langs[0]
        #     others = record.langs[1:]
        #     lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # else:
        #     lang_list = f"$a{"$a".join(record.lang)}"
        ## OPTION 2
        # main_lang = record.langs[0]
        # others = record.langs[1:]
        # lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # result = (build_field(41, [[i1,i2, lang_list]]))
        ## OPTION 3
        result = Result(Field(tag, i1, i2, [Subfield("a", language) for language in record.langs]), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_246(record: Record) -> Result:  ##optional
    """
    Varying Form of Title
    Holds parallel Western title AND/OR original Chinese character title
    NB: Initial articles (e.g., The, La) are generally not recorded in field 246 unless the intent is to file on the article. [https://www.loc.gov/marc/bibliographic/bd246.html]
    """
    tag = 246
    i1 = 3  ## "No note, added entry"
    i2 = 1  ## parallel title
    has_parallel_title = bool(record.parallel_title.original)
    has_chinese_parallel_title = bool(record.parallel_title.transliteration)
    linkage: Subfield | None = None
    if has_chinese_parallel_title:
        parallel_title = record.parallel_title.transliteration
        parallel_subtitle = record.parallel_subtitle.transliteration
        linkage = deal_with_chinese_titles(record, record.parallel_title.original, record.parallel_subtitle.original, i1, i2, tag)
    elif has_parallel_title:  ## (i.e. Western script)
        parallel_title, parallel_subtitle = record.parallel_title.original, record.parallel_subtitle.original
    else:
        parallel_title, parallel_subtitle = "", ""
    if parallel_title:
        parallel_title_sub = Subfield("a", parallel_title)
        if linkage:
           content = Content([linkage, parallel_title_sub])
        else:
            content = Content(parallel_title_sub)
        if parallel_subtitle:
            content.add(Subfield("b", parallel_subtitle))
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_490(record: Record) -> Result:  ## optional
    """Series Statement"""
    tag = 490
    i1 = 0  ## Series not traced: No series added entry is desired for the series.
    i2 = -1
    series_title = Subfield("a", record.series_title if record.series_title else "")
    series_enum = Subfield("v", record.series_enum if record.series_enum else "")
    content: Content | None = None
    if record.series_title and record.series_enum:
        content = Content([series_title, Punctuation(" ;"), series_enum])
    elif record.series_title:
        content = Content(series_title)
    elif record.series_enum:
        content = Content(series_enum)
    if content:
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_500(record: Record) -> Result:  ##optional
    """general notes
    Punctuation - Field 500 ends with a period unless another mark of punctuation is present. If the final subfield is subfield $5, the mark of punctuation precedes that subfield.
    """
    tag = 500
    i1 = -1
    i2 = -1
    if record.notes:
        content = Content(Subfield("a", record.notes))
        if content.can_accept_period():
            content.add(Punctuation("."))
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


# TODO: need an item with both a Chinese title and subtitle to test the sequence number logic.
def build_880(record: Record, title: Content, i1: int, i2: int, caller: int, sequence_number: str) -> None:  ##optional
    """Alternate Graphic Representation
    NB. unlike the other fields, this isn't called directly but by the linked field
    looks like: =880  31$6246-01$a中國書畫、陶瓷及藝術品拍賣會
    """
    record.sequence_number += 1
    content = Content(Subfield(f"{str(caller).zfill(3)}-{sequence_number}", 6))
    content.add(title.contents)
    line = Field(880, i1, i2, content)
    record.links.append(line)


def check_if_mandatory(result: Result, is_mandatory: bool) -> list[Field] | None:
    """
    silently suppresses optional fields if empty;
    stops with error if required field is empty
    """
    output: list[Field] | None = None
    if result.is_err:
        numeric_tag, error_msg = result.is_err
        if is_mandatory:
            msg = error_msg if error_msg else f"Data for required field {str(numeric_tag).zfill(3)} is required."
            logger.warning(msg)
            raise MissingFieldError(msg)
    elif result.is_ok:
        if isinstance(result.is_ok, list):
            output = result.is_ok
        else:
            output = [result.is_ok]
    return output


def line_prefix(numeric_tag: int) -> str:
    display_tag = "LDR" if numeric_tag == 0 else seq_num(numeric_tag)
    return f"={expand_tag(display_tag)}  "


def expand_tag(field_number: str) -> str:
    return str(field_number).zfill(3)


def seq_num(sequence_number: int) -> str:
    return str(sequence_number).zfill(2)


def check_for_nonfiling(title: str, lang: str="eng") -> tuple[int,str]:
    """
    Check for manual nonfiling indicator (@@) & returns its position + extracts it from title string
    if no manual indication, check for nonfiling words according to language of title"""
    nonfiling_words = {
        "eng": ("the", "a", "an"),
        "fr": ("le", "la", "les", "l'", "un", "une"),
        "it": ("lo", "il", "i", "l'", "gli", "le", "un", "una", "un'"),
        "sp": ("el", "la", "las", "los", "un", "una", "unos", "unas"),
        "gw": ("der", "die", "das", "ein", "eine"),
        "ne": ("de", "het"),
        "sw": ("en", "ett", "den", "det", "de"),
        "dk": ("en", "et"),
        "no": ("en", "ei", "et"),
        "po": ("o", "a", "os", "as", "um", "uma", "uns", "umas"),
        "chi": (),
    }
    break_char = "@@"
    nonfiling = title.find(break_char)
    result = (0, title)
    if nonfiling > 0:
        result = (nonfiling, title.replace(break_char, "", 1))
    else:
        for article in nonfiling_words[lang]:
            test = re.match(f"({article}\\s?[^\\w\\s]?)\\w", title, re.I)
            if test:
                result = (test.span()[1] - 1, title)
                break
    return result


def variable_control_field():
    return (-2, -2)


def build_marc_records(records: list[Record]) -> list[marc_record]:
    # marc_records: list[list[Field]] = []
    marc_records: list[marc_record] = []
    for record in records:
        # print(record, type(record))
        # pprint(record)
        marc = apply_marc_logic(record)
        marc_records.append(marc)
    return marc_records


def apply_marc_logic(record: Record) -> list[Field]:
    marc_record: list[Field] = []
    # print(record, type(record))
    # fields_to_deploy: tuple[tuple[Callable[[Record], Result], bool]] = (
    fields_to_deploy: tuple[tuple[Callable, bool]] = (
        (build_leader, True),
        (build_040, True),
        (build_336, True),
        (build_337, True),
        (build_338, True),
        (build_904, True),
        (build_005, True),
        (build_008, True),
        (build_033, True),
        (build_245, True),
        (build_264, True),
        (build_300, True),
        (build_490, False), # series statement
        (build_876, True),
        (build_020, False), # isbn
        (build_024, False), # sales code
        (build_041, False), # language if not monolingual
        (build_246, False), # parallel title 
        (build_500, False), # general notes
    )
    for builder, is_mandatory in fields_to_deploy:
        field = check_if_mandatory(builder(record), is_mandatory)
        if field:
            for repeat in field:
                marc_record.append(repeat)
    if len(record.links):
        for link in record.links:
            if link:
                marc_record.append(link)
    marc_record.sort(key=lambda x: (x.tag * 10) + x.ordering)
    return marc_record


def write_mrk_files(data: list[list[str]], file_name: str="output.mrk") -> None:
    mrk_file_dir = make_directory("marc21_files")
    out_file = mrk_file_dir / file_name
    with open(out_file, "w", encoding="utf-8") as f:
        for record in data:
            for field in record:
                f.write(field + "\n")
            f.write("\n")


def write_mrc_binaries(data: list[list[Field]], file_name: str="output.mrc") -> None:
    mrc_file_dir = make_directory("marc21_files")
    out_file = mrc_file_dir / file_name
    flat_records = make_binary(data)
    # print(flat_records)
    # with open(out_file, "w", encoding="utf-8") as f:
    with open(out_file, "wb") as f:
        for line in flat_records:
            # f.write(line.encode("utf-8"))
            f.write(line)


def make_directory(directory_path):
    directory = Path(directory_path)
    if not directory.is_dir():
        directory.mkdir()
        try:
            directory.mkdir()
            logger.info(f"Directory '{directory}' created successfully.")
        except FileExistsError:
            logger.info(f"Directory '{directory}' already exists.")
        except PermissionError:
            logger.warning(f"Permission denied: Unable to create '{directory}'.")
        except Exception as e:
            logger.warning(f"An error occurred: {e}")
    return directory


# def make_binary(data: list[list[Field]]) -> list[str]:
def make_binary(data: list[list[Field]]) -> list[bytes]:
    # GS = "@"
    # RS = "%"
    # US = "#"
    GS = chr(29)
    RS = chr(30)
    # output: list[str] = []
    output: list[bytes] = []
    tmp = []
    for record in data:
        leader: Field = record[0]
        fields = []
        current_start_position = 0
        for field in record[1:]:
            tag = field.tag
            contents = field.to_mrc()
            line_length_with_final_record_separator = len(contents.encode("utf-8"))
            fields.append((tag, contents, line_length_with_final_record_separator, current_start_position))
            current_start_position += line_length_with_final_record_separator
        tmp.append((leader.to_mrc(), fields))

        directory = [(f[0], f[2], f[3]) for f in fields]
        leader_length = int(24)
        directory_length = (12 * len(directory)) + 1
        base_address_of_data = leader_length + directory_length
        logical_record_length = base_address_of_data + directory[-1][2]
        leader_mrc = f"{str(logical_record_length).zfill(5)}nam a22{str(base_address_of_data).zfill(5)}3i 4500"
        directory_mrc = "".join([f"{str(f[0]).zfill(3)}{str(f[1]).zfill(4)}{str(f[2]).zfill(5)}" for f in directory]) + RS
        fields_mrc = "".join([f[1] for f in fields])
        # output.append(f"{leader_mrc}{directory_mrc}{fields_mrc}{GS}")
        binary_line = f"{leader_mrc}{directory_mrc}{fields_mrc}{GS}".encode("utf-8")
        output.append(binary_line)
    return output


def parse_excel_into_rows(excel_file_address: Path) -> list[worksheet_row]:
    excel_file_name = str(excel_file_address.resolve())
    worksheet = load_workbook(filename=excel_file_name).active
    raw_rows = extract_from_excel(worksheet)
    return raw_rows


def flatten_fields_to_strings(input: list[marc_record]) -> list[list[str]]:
    output: list[list[str]] = []
    for i, record in enumerate(input):
        # print(f">>>>>>>>>>{i} -> {record}")
        output.append([field.to_string() for field in record])
    return output


def write_marc_files(records:list[marc_record], excel_file_address: Path) -> None:
    records_with_string_fields = flatten_fields_to_strings(records)
    # print(records_with_string_fields)
    # pprint(records_with_marc_fields)
    write_mrk_files(records_with_string_fields, f"{excel_file_address.stem}.paul.mrk")
    write_mrc_binaries(records, f"{excel_file_address.stem}.paul.mrc")


def run() -> None:
    # process_excel_file(sys.argv[1])
    # process_excel_file(Path("excel_files") / "chinese_test.xlsx")
    # quit()
    # for file in Path("excel_files").glob("*.xls?"):
    for file in Path("excel_files").glob("*.xlsx"):
        logger.info(f"\n>>>>> processing: {file.name}")
        print(f">>>>> processing: {file.name}")
        raw_rows = parse_excel_into_rows(file)
        records = parse_rows_into_records(raw_rows)
        del raw_rows
        marc_records = build_marc_records(records)
        del records
        write_marc_files(marc_records, file)
        # make_marc_files(raw_records, file)


def main() -> None:
    run()

if __name__ == "__main__":
    logger = logging.getLogger(__name__)
    main()