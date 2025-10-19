from openpyxl import load_workbook  # type: ignore
import csv
from dataclasses import dataclass, fields, field
from abc import ABC, abstractmethod
from typing import TypeAlias, Any
from collections.abc import Callable

# from pprint import pprint
# import sys
from datetime import datetime, timezone
import re
from pathlib import Path
import logging

# from pyisbn import Isbn
# from pydantic import BaseModel, field_validator, ValidationError
# from pydantic_extra_types.isbn import ISBN


def make_directory(directory_path:str) -> Path:
    directory = Path(directory_path)
    if not directory.is_dir():
        directory.mkdir()
        try:
            directory.mkdir()
            logger.info(f"Directory '{directory}' created successfully.")
        except FileExistsError:
            logger.info(f"Directory '{directory}' already exists.")
        except PermissionError:
            logger.warning(f"Permission denied: Unable to create '{directory}'.")
        except Exception as e:
            logger.warning(f"An error occurred: {e}")
    return directory


@dataclass
class Settings:
    in_file_full = ""
    in_file = ""
    out_file = ""
    default_output_filename = "output"
    # data_dir = "excel_files"
    # output_dir = "marc21_files"
    app_dir = "app"
    data_dir = "input_files"
    output_dir = "output_files"
    use_default_layout = True
    is_existing_file = True
    layout_template: tuple = ()
    first_row_is_header = True
    flavour: dict[str, Any] = field(default_factory=dict)
    styles: dict[str, str]= field(default_factory=dict)
    help_file = ""
    backup_file = "backup.bak"


logger = logging.getLogger(__name__)
settings = Settings()
excel_file_dir = make_directory(settings.data_dir)
output_file_dir = make_directory(settings.output_dir)


logging.basicConfig(
    filename=output_file_dir / "output.log",
    filemode="w",
    encoding="utf-8",
    format="%(levelname)s:%(message)s",
    level=logging.DEBUG,
)


class Serializable(ABC):
    @abstractmethod
    def serialize(self, mode="str") -> str:
        pass


@dataclass
class Blank(Serializable):
    def serialize(self, mode="str") -> str:
        if mode == "str":
            return "\\"
        else:
            return " "


class Punctuation(Serializable):
    """ISBD punctuation used between subfields
    https://www.itsmarc.com/crs/mergedprojects/lcri/lcri/1_0c__lcri.htm
    """

    def __init__(self, contents: str):
        self.contents = contents

    def serialize(self, mode="str") -> str:
        return self.contents

    def __str__(self):
        return f"<contents='{self.contents}'>"


class Subfield(Serializable):
    """Subfield class for MARC21 subfields.
    -1 means the content is a single string (for variable control fields)
    if code is omitted, assumed to be -1 (i.e. no code)
    """

    # def __init__(self, code: str | int, data: str):
    def __init__(self, data: str, code: str | int = "-1"):
        self.code: str = code if isinstance(code, str) else str(code)
        self.data: str = data

    def serialize(self, mode="str") -> str:
        delimiter = "$" if mode == "str" else chr(31)
        if self.code == "-1":
            prefix = ""
        else:
            prefix = f"{delimiter}{self.code}"
        return f"{prefix}{self.data}"

    def __repl__(self):
        return f"<code='{self.code}', data='{self.data}'>"


atoms: TypeAlias = Subfield | Punctuation | Blank
worksheet_row: TypeAlias = list[str]


class Content(Serializable):
    def __init__(self, contents: atoms | list[atoms]):
        self.contents = []
        if not isinstance(contents, list):
            contents = [contents]
        for el in contents:
            if type(el) not in [Subfield, Punctuation, Blank]:
                raise TypeError("Only subfields or punctuation allowed")
            else:
                self.contents.append(el)

    def can_accept_period(self) -> bool:
        return bool(self.serialize().rstrip()[-1] not in "?!.")

    def serialize(self, mode="str") -> str:
        return "".join([el.serialize(mode) for el in self.contents])

    def add(self, content: list[atoms] | atoms):
        if isinstance(content, list):
            self.contents.extend(content)
        else:
            self.contents.append(content)

    def __getitem__(self, index: int):
        return self.contents[index]

    def __str__(self):
        result = ", ".join([str(el) for el in self.contents])
        return f"[{result}]"


class Field(Serializable):
    """Variable Control/Data Field class for MARC21 variable control fields.
    -1 or Blank = indicator placeholder (various realizations, e.g. space or backslash)
    -2 = no indicator"""

    def __init__(
        self,
        tag: int,
        i1: int | Blank,
        i2: int | Blank,
        contents: Content | atoms | list[atoms],
        ordering: int = 1,
    ):
        self.tag = tag
        self.i1: int | Blank = Blank() if i1 == -1 else i1
        self.i2: int | Blank = Blank() if i2 == -1 else i2
        self.contents = contents if isinstance(contents, Content) else Content(contents)
        self.ordering: int = (
            ordering  ## sort is used to order the fields when there are more than one of the same tag
        )

    def expand_indicators(self, indicator: int | Blank, mode="str") -> str:
        if indicator == -1:
            indicator = Blank()
        if isinstance(indicator, Blank):
            expansion = indicator.serialize(mode)
        elif indicator == -2:
            expansion = ""
        else:
            expansion = str(indicator)
        return expansion

    def can_accept_period(self) -> bool:
        return self.contents.can_accept_period()

    def serialize(self, mode="str") -> str:
        i1 = self.expand_indicators(self.i1, mode)
        i2 = self.expand_indicators(self.i2, mode)
        contents = self.contents.serialize(mode)
        tag, end = "", ""
        if mode == "str":
            tag = "LDR" if self.tag == 0 else str(self.tag).zfill(3)
            tag = f"={tag}  "
        else:
            end = chr(30)
        return f"{tag}{i1}{i2}{contents}{end}"

    def __repr__(self):
        # return f"={self.tag}@{self.ordering} ({self.i1})({self.i2}) {self.contents.to_string()}"
        return f"={self.tag}@{self.ordering} ({self.i1})({self.i2}) {self.contents.serialize()}"


marc_record: TypeAlias = list[Field]


@dataclass
class Title:
    original: str
    transliteration: str


@dataclass
class Record:
    sublib: str
    langs: list[str]
    isbn: str
    title: Title
    subtitle: Title
    parallel_title: Title
    parallel_subtitle: Title
    country_name: str
    country_code: str
    state: str
    place: str
    publisher: str
    pub_year: str
    copyright: str
    extent: str
    size: int
    series_title: str
    series_enum: str
    volume: str
    notes: str
    sales_code: str
    sale_dates: list[str]
    hol_notes: str
    donation: str
    barcode: str

    pub_year_is_approx: bool
    extent_is_approx: bool
    timestamp: datetime

    sequence_number: int
    links: list[Field | None]


@dataclass
class Result:
    """
    holds the successfully acquired data OR
    the MARC field tag (as an int) where the problem occurred and an error message
    """

    is_ok: list[Field] | Field | None  ## MARC field or list of same
    is_err: tuple[int, str] | None  ## [field number, error message]


class MissingFieldError(Exception):
    pass


def norm_langs(raw: str) -> list[str]:
    language_codes = {
        "english": "eng",
        "chinese": "chi",
        "german": "ger",
        "italian": "ita",
        "spanish": "spa",
        "french": "fre",
        "swedish": "swe",
        "danish": "dan",
        "norwegian": "nor",
        "dutch": "dut",
    }
    list_of_languages = []
    languages: list[str] = re.split("[,/]", raw.replace(" ", "").lower())
    for language in languages:
        try:
            list_of_languages.append(language_codes[language])
        except KeyError as e:
            logger.warning(
                f"Warning: {e} is not a recognised language; it has been passed on unchanged."
            )
            list_of_languages.append(language)
    return list_of_languages


def norm_geographical_name(name: str) -> str:
    return re.sub(r"[\s\-\.']", "", name).lower()


def norm_country(country_raw: str) -> str:
    country_codes = {
        "usa": "xxu",
        "unitedstates": "xxu",
        "unitedstatesofamerica": "xxu",
        "uk": "xxk",
        "unitedkingdom": "xxk",
        "canada": "xxc",
        "australia": "xxa",  # dummy 3-digit value to trigger detail; swap to "at" at output
        "algeria": "ae",
        "angola": "ao",
        "benin": "dm",
        "botswana": "bs",
        "burkinafaso": "uv",
        "burundi": "bd",
        "cameroon": "cm",
        "centralafricanrepublic": "cx",
        "chad": "cd",
        "congo": "cf",
        "democraticrepublicofcongo": "cg",
        "côtedivoire": "iv",
        "cotedivoire": "iv",
        "djibouti": "ft",
        "egypt": "ua",
        "equatorialguinea": "eg",
        "eritrea": "ea",
        "ethiopia": "et",
        "gabon": "go",
        "gambia": "gm",
        "ghana": "gh",
        "guinea": "gv",
        "guineabissau": "pg",
        "kenya": "ke",
        "lesotho": "lo",
        "liberia": "lb",
        "libya": "ly",
        "madagascar": "mg",
        "malawi": "mw",
        "mali": "ml",
        "mauritania": "mu",
        "morocco": "mr",
        "mozambique": "mz",
        "namibia": "sx",
        "niger": "ng",
        "nigeria": "nr",
        "rwanda": "rw",
        "saotomeandprincipe": "sf",
        "senegal": "sg",
        "sierraleone": "sl",
        "somalia": "so",
        "southafrica": "sa",
        "southsudan": "sd",
        "spanishnorthafrica": "sh",
        "sudan": "sj",
        "swaziland": "sq",
        "tanzania": "tz",
        "togo": "tg",
        "tunisia": "ti",
        "uganda": "ug",
        "westernsahara": "ss",
        "zambia": "za",
        "zimbabwe": "rh",
        "afghanistan": "af",
        "armenia": "ai",
        "republicofarmenia": "ar",
        "azerbaijan": "aj",
        "bahrain": "ba",
        "bangladesh": "bg",
        "bhutan": "bt",
        "brunei": "bx",
        "burma": "br",
        "cambodia": "cb",
        "china": "cc",
        "cyprus": "cy",
        "easttimor": "em",
        "gazastrip": "gz",
        "georgia": "gs",
        "georgianrepublic": "gs",
        "republicofgeorgia": "gs",
        "india": "ii",
        "indonesia": "io",
        "iran": "ir",
        "iraq": "iq",
        "israel": "is",
        "japan": "ja",
        "jordan": "jo",
        "kazakhstan": "kz",
        "northkorea": "kn",
        "korea": "ko",
        "southkorea": "ko",
        "kuwait": "ku",
        "kyrgyzstan": "kg",
        "laos": "ls",
        "lebanon": "le",
        "malaysia": "my",
        "mongolia": "mp",
        "nepal": "np",
        "oman": "mk",
        "pakistan": "pk",
        "papuanewguinea": "pp",
        "paracelislands": "pf",
        "philippines": "ph",
        "qatar": "qa",
        "saudiarabia": "su",
        "singapore": "si",
        "spratlyisland": "xp",
        "srilanka": "ce",
        "syria": "sy",
        "tajikistan": "ta",
        "thailand": "th",
        "turkey": "tu",
        "turkmenistan": "tk",
        "unitedarabemirates": "ts",
        "uae": "ts",
        "uzbekistan": "uz",
        "vietnam": "vm",
        "westbankofthejordanriver": "wj",
        "westbank": "wj",
        "yemen": "ye",
        "bermudaislands": "bm",
        "bermuda": "bm",
        "bouvetisland": "bv",
        "caboverde": "cv",
        "faroeislands": "fa",
        "faroes": "fa",
        "falklandislands": "fk",
        "falklands": "fk",
        "sainthelena": "xj",
        "southgeorgiaandthesouthsandwichislands": "xs",
        "southgeorgia": "xs",
        "southsandwichislands": "xs",
        "belize": "bh",
        "costarica": "cr",
        "elsalvador": "es",
        "guatemala": "gt",
        "honduras": "ho",
        "nicaragua": "nq",
        "panama": "pn",
        "albania": "aa",
        "andorra": "an",
        "austria": "au",
        "belarus": "bw",
        "belgium": "be",
        "bosniaandherzegovina": "bn",
        "bosnia": "bn",
        "bosniaherzegovina": "bn",
        "herzegovina": "bn",
        "bulgaria": "bu",
        "croatia": "ci",
        "czechrepublic": "xr",
        "czechia": "xr",
        "denmark": "dk",
        "estonia": "er",
        "finland": "fi",
        "france": "fr",
        "germany": "gw",
        "gibraltar": "gi",
        "greece": "gr",
        "guernsey": "gg",
        "hungary": "hu",
        "iceland": "ic",
        "ireland": "ie",
        "isleofman": "im",
        "italy": "it",
        "jersey": "je",
        "kosovo": "kv",
        "latvia": "lv",
        "liechtenstein": "lh",
        "lithuania": "li",
        "luxembourg": "lu",
        "macedonia": "xn",
        "malta": "mm",
        "montenegro": "mo",
        "moldova": "mv",
        "monaco": "mc",
        "netherlands": "ne",
        "norway": "no",
        "poland": "pl",
        "portugal": "po",
        "serbia": "rb",
        "romania": "rm",
        "russia": "ru",
        "russianfederation": "ru",
        "sanmarino": "sm",
        "slovakia": "xo",
        "slovenia": "xv",
        "spain": "sp",
        "sweden": "sw",
        "switzerland": "sz",
        "ukraine": "un",
        "vaticancity": "vc",
        "serbiaandmontenegro": "yu",
        "britishindianoceanterritory": "bi",
        "christmasisland": "xa",
        "cocosislands": "xb",
        "keelingislands": "xb",
        "comoros": "cq",
        "heardandmcdonaldislands": "hm",
        "maldives": "xc",
        "mauritius": "mf",
        "mayotte": "ot",
        "réunion": "re",
        "reunion": "re",
        "seychelles": "se",
        "americansamoa": "as",
        "cookislands": "cw",
        "fiji": "fj",
        "frenchpolynesia": "fp",
        "guam": "gu",
        "johnstonatoll": "ji",
        "kiribati": "gb",
        "marshallislands": "xe",
        "micronesia": "fm",
        "federatedstatesofmicronesia": "fm",
        "midwayislands": "xf",
        "nauru": "nu",
        "newcaledonia": "nl",
        "niue": "xh",
        "northernmarianaislands": "nw",
        "palau": "pw",
        "pitcairnisland": "pc",
        "samoa": "ws",
        "solomonislands": "bp",
        "tokelau": "tl",
        "tonga": "to",
        "tuvalu": "tv",
        "vanuatu": "nn",
        "wakeisland": "wk",
        "wallisandfutuna": "wf",
        "wallis": "wf",
        "futuna": "wf",
        "argentina": "ag",
        "bolivia": "bo",
        "brazil": "bl",
        "chile": "cl",
        "colombia": "ck",
        "ecuador": "ec",
        "frenchguiana": "fg",
        "guyana": "gy",
        "paraguay": "py",
        "peru": "pe",
        "surinam": "sr",
        "uruguay": "uy",
        "venezuela": "ve",
        "anguilla": "am",
        "antiguaandbarbuda": "aq",
        "antigua": "aq",
        "barbuda": "aq",
        "aruba": "aw",
        "bahamas": "bf",
        "barbados": "bb",
        "britishvirginislands": "vb",
        "caribbeannetherlands": "ca",
        "caymanislands": "cj",
        "cuba": "cu",
        "curaçao": "co",
        "curacao": "co",
        "dominica": "dq",
        "dominicanrepublic": "dr",
        "grenada": "gd",
        "guadeloupe": "gp",
        "haiti": "ht",
        "jamaica": "jm",
        "martinique": "mq",
        "montserrat": "mj",
        "puertorico": "pr",
        "saintbarthélemy": "sc",
        "saintbarthelemy": "sc",
        "saintkittsnevis": "xd",
        "saintkitts": "xd",
        "nevis": "xd",
        "saintlucia": "xk",
        "saintmartin": "st",
        "saintvincentandthegrenadines": "xm",
        "saintvincent": "xm",
        "thegrenadines": "xm",
        "grenadines": "xm",
        "sintmaarten": "sn",
        "trinidadandtobago": "tr",
        "trinidad": "tr",
        "tobago": "tr",
        "turksandcaicosislands": "tc",
        "virginislandsoftheunitedstates": "vi",
        "antarctica": "ay",
        "noplace": "xx",
        "unknown": "xx",
        "undetermined": "xx",
        "variousplaces": "vp",
        "various": "vp",
    }
    country = norm_geographical_name(country_raw)
    try:
        result = country_codes[country]
    except KeyError as e:
        if len({e}) < 4:
            logger.info(
                f"Advisory: assuming country name ({e}) has already been processed."
            )
        else:
            logger.warning(
                f"Warning: {e} is not a recognised country name; it has been passed on unchanged."
            )
        result = country_raw
    return result


def norm_place(place_raw: str) -> str:
    long_country_codes = {
        "england": "enk",
        "northernireland": "nik",
        "scotland": "stk",
        "wales": "wlk",
        "alberta": "abc",
        "britishcolumbia": "bcc",
        "manitoba": "mbc",
        "newbrunswick": "nkc",
        "newfoundland": "nfc",
        "labrador": "nfc",
        "newfoundlandandlabrador": "nfc",
        "northwestterritories": "ntc",
        "novascotia": "nsc",
        "nunavut": "nuc",
        "ontario": "onc",
        "princeedwardisland": "pic",
        "québecprovince": "quc",
        "québec": "quc",
        "quebecprovince": "quc",
        "quebec": "quc",
        "saskatchewan": "snc",
        "yukonterritory": "ykc",
        "yukon": "ykc",
        "alabama": "alu",
        "alaska": "aku",
        "arizona": "azu",
        "arkansas": "aru",
        "california": "cau",
        "colorado": "cou",
        "connecticut": "ctu",
        "delaware": "deu",
        "districtofcolumbia": "dcu",
        "columbia": "dcu",
        "florida": "flu",
        "georgia": "gau",
        "hawaii": "hiu",
        "idaho": "idu",
        "illinois": "ilu",
        "indiana": "inu",
        "iowa": "iau",
        "kansas": "ksu",
        "kentucky": "kyu",
        "louisiana": "lau",
        "maine": "meu",
        "maryland": "mdu",
        "massachusetts": "mau",
        "michigan": "miu",
        "minnesota": "mnu",
        "mississippi": "msu",
        "missouri": "mou",
        "montana": "mtu",
        "nebraska": "nbu",
        "nevada": "nvu",
        "newhampshire": "nhu",
        "newjersey": "nju",
        "newmexico": "nmu",
        "newyork": "nyu",
        "newyorkstate": "nyu",
        "northcarolina": "ncu",
        "northdakota": "ndu",
        "ohio": "ohu",
        "oklahoma": "oku",
        "oregon": "oru",
        "pennsylvania": "pau",
        "rhodeisland": "riu",
        "southcarolina": "scu",
        "southdakota": "sdu",
        "tennessee": "tnu",
        "texas": "txu",
        "utah": "utu",
        "vermont": "vtu",
        "virginia": "vau",
        "washington": "wau",
        "washingtonstate": "wau",
        "westvirginia": "wvu",
        "wisconsin": "wiu",
        "wyoming": "wyu",
        "australiancapitalterritory": "aca",
        "queensland": "qea",
        "tasmania": "tma",
        "victoria": "vra",
        "westernaustralia": "wea",
        "newsouthwales": "xna",
        "northernterritory": "xoa",
        "southaustralia": "xra",
        "al": "alu",
        "ak": "aku",
        "az": "azu",
        "ar": "aru",
        "ca": "cau",
        "co": "cou",
        "ct": "ctu",
        "de": "deu",
        "dc": "dcu",
        "fl": "flu",
        "ga": "gau",
        "hi": "hiu",
        "id": "idu",
        "il": "ilu",
        "in": "inu",
        "ia": "iau",
        "ks": "ksu",
        "ky": "kyu",
        "la": "lau",
        "me": "meu",
        "md": "mdu",
        "ma": "mau",
        "mi": "miu",
        "mn": "mnu",
        "ms": "msu",
        "mo": "mou",
        "mt": "mtu",
        "ne": "nbu",
        "nv": "nvu",
        "nh": "nhu",
        "nj": "nju",
        "nm": "nmu",
        "ny": "nyu",
        "nc": "ncu",
        "nd": "ndu",
        "oh": "ohu",
        "ok": "oku",
        "or": "oru",
        "pa": "pau",
        "ri": "riu",
        "sc": "scu",
        "sd": "sdu",
        "tn": "tnu",
        "tx": "txu",
        "ut": "utu",
        "vt": "vtu",
        "va": "vau",
        "wa": "wau",
        "wv": "wvu",
        "wi": "wiu",
        "wy": "wyu",
        "act": "aca",
        "qld": "qea",
        "tas": "tma",
        "vic": "vra",
        "wa": "wea",
        "nsw": "xna",
        "nt": "xoa",
        "sa": "xra",
        "alb": "abc",
        "bc": "bcc",
        "man": "mbc",
        "nb": "nkc",
        "nfd": "nfc",
        "lab": "nfc",
        "nwt": "ntc",
        "ns": "nsc",
        "nu": "nuc",
        "ont": "onc",
        "pei": "pic",
        "que": "quc",
        "sas": "snc",
        "yt": "ykc",
    }
    place = norm_geographical_name(place_raw)
    try:
        result = long_country_codes[place]
    except KeyError as e:
        if len({e}) == 3:
            logger.info(
                f"Advisory: assuming place name ({e}) has already been processed."
            )
        else:
            logger.warning(
                f"Warning: {e} is not a recognised place name; it has been passed on unchanged."
            )
        result = place_raw
    return result


def norm_state(state_raw: str) -> str:
    if state_raw:
        return norm_place(state_raw)
    else:
        return ""


def check_for_detailed_region(country: str, state: str, place: str) -> str:
    """
    The country code for USA, Australia, Canada & UK is 3 digit, based on the state
    Australia, however, only has a 2 digit superordinate code ('at')
    """
    region = country
    if len(region) == 3:
        if state:
            region = norm_state(state)
        else:
            tmp = norm_state(place)  # maybe state entered as city by mistake
            if len(tmp) == 3:
                region = tmp
        if (
            region == "xxa"
        ):  # unlike other 3-digit regions, australia only has 2 digits :S
            region = "at"
        # print(f"**** {country}, {state} ({len(state)}), {place}  -> {region}")
    return region


def validate_record(record: Record) -> bool:
    mandatory = [
        "sublib",
        "langs",
        "title",
        "country",
        # "place",
        "publisher",
        "pub_year",
        "extent",
        "size",
        "sale_dates",
        "barcode",
    ]
    is_valid = False
    place_state_check = 0
    for i, field in enumerate(fields(record)):
        name = field.name
        is_valid = True
        if name in ("place", "state"):
            place_state_check += 1
        if name in mandatory and not getattr(record, name):
            logger.warning(f"Record no. {i} is missing the mandatory field '{name}'.")
            is_valid = False
    if not place_state_check:
        logger.warning(
            "A record must have EITHER a place OR a state specified; this lacks both."
        )
    return is_valid


def norm_dates(raw: str) -> list[str]:
    result = [date.strip() for date in raw.split(",")]
    return result


def norm_size(raw: str) -> int:
    raw = strip_unwanted(r"cm", raw)
    clean_value = int(raw) if raw.isnumeric() else -1
    return clean_value


def norm_pages(pages_raw: str) -> str:
    pages = strip_unwanted(r"pages|\[|\]", pages_raw)
    if "approx" in pages:
        pages = re.sub(r"[^\d]", "", pages)
        pages = pages + "?"
    return pages


def norm_year(year_raw: str) -> str:
    year = strip_unwanted(r"[\[\]]", year_raw)
    return year


def norm_isbn(raw_isbn: str) -> str:
    isbn = re.sub(r"[\s-]", "", raw_isbn)
    if 10 > len(isbn) > 13:
        msg = f"isbn {raw_isbn} is non-standard"
        print(f"*** {msg}")
        logger.warning(msg)
    return isbn


def norm_barcode(raw_barcode: str) -> str:
    if not re.match(r"^[367]\d{8}", raw_barcode):
        msg = f"barcode {raw_barcode} is non-standard"
        print(f"*** {msg}")
        logger.warning(msg)
    return raw_barcode


def strip_unwanted(pattern: str, raw: str) -> str:
    clean = re.sub(pattern, "", raw).strip()
    return clean


def check_for_approx(raw_string: str) -> tuple[str, bool]:
    """
    take a question mark in any position to mean the information is not certain.
    """
    clean, is_approx = re.subn(r"\?", "", raw_string)
    clean = trim_mistaken_decimals(clean).strip()
    return (clean, bool(is_approx))


def trim_mistaken_decimals(value: str | int) -> str:
    if not isinstance(value, str):
        value = str(value)
    if value.endswith(".0"):
        value = value[:-2]
    return value


# TODO: what should this return if nothing? None or []?
def fill_with_blanks(string: str, limit: int = 3) -> list[Blank]:
    return [Blank() for _ in range(limit - len(string))]


def create_date_list(dates_raw: str) -> list[str]:
    dates_raw = re.sub(r"\s|\.0", "", dates_raw)
    dates = dates_raw.split(",")
    return dates


def extract_from_excel(excel_sheet) -> tuple[list[str], list[worksheet_row]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    sheet = []
    headers = []
    # for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
    for i, excel_row in enumerate(excel_sheet.iter_rows(min_row=1, values_only=True)):
        if not excel_row[0] and not excel_row[1]:
            break  ## needed as openpyxl keeps spitting out empty rows at the end
        row = normalize_row(excel_row)
        if i == 0 and settings.first_row_is_header:
            headers = row
        else:
            sheet.append(row)
    return (headers, sheet)


def normalize_row(row: list) -> list:
    clean_row = []
    for col in row:
        if col:
            data = str(col).strip()
            data = trim_mistaken_decimals(data)
        else:
            data = ""
        clean_row.append(data)
    return clean_row


def parse_rows_into_records(sheet: list[worksheet_row]) -> list[Record]:
    current_time = datetime.now()
    records = []
    for row in sheet:
        # for i, row in enumerate(sheet):
        #     print(f"Parsing record no. {i + 1}")
        record = parse_row(row, current_time)
        records.append(record)
    return records


def parse_row(row: list[str], current_time: datetime) -> Record:
    cols = iter(row)
    sublibrary = next(cols)
    langs = norm_langs(next(cols))
    isbn = norm_isbn(next(cols))
    title = Title(next(cols), next(cols))
    subtitle = Title(next(cols), next(cols))
    parallel_title = Title(next(cols), next(cols))
    parallel_subtitle = Title(next(cols), next(cols))
    country_name = next(cols)
    country = norm_country(country_name)
    state = next(cols)
    place = next(cols)
    publisher = next(cols)
    pub_date, pub_date_is_approx = check_for_approx(norm_year(next(cols)))
    copyright_ = next(cols).replace("©", "").strip()
    extent, extent_is_approx = check_for_approx(norm_pages(next(cols)))
    size = norm_size(next(cols))
    series_title = next(cols)
    series_enum = next(cols)
    volume = next(cols)
    note = next(cols)
    sale_code = next(cols)
    date_of_sale = create_date_list(next(cols))
    hol_notes = next(cols)
    donation = next(cols)
    barcode = norm_barcode(next(cols))

    record = Record(
        sublibrary,
        langs,
        isbn,
        title,
        subtitle,
        parallel_title,
        parallel_subtitle,
        country_name,
        country,
        state,
        place,
        publisher,
        pub_date,
        copyright_,
        extent,
        size,
        series_title,
        series_enum,
        volume,
        note,
        sale_code,
        date_of_sale,
        hol_notes,
        donation,
        barcode,
        pub_date_is_approx,
        extent_is_approx,
        current_time,
        sequence_number=1,
        links=[],
    )
    validate_record(record)
    return record


def build_leader(record: Record) -> Result:
    """leader (0 is only for sorting purposes; should read 'LDR')"""
    tag = 0
    blank = " "
    i1, i2 = variable_control_field()
    # content = "00000nam a22000003i 4500"
    record_len_00 = "00000"  # placeholder for record length
    record_status_05 = "n"  # "n" for new record
    record_type_06 = "a"  # "a" for language material
    biblio_level_07 = "m"  # "m" for monograph
    type_of_ctrl_08 = Blank()  # no type of control
    char_coding_09 = "a"  # Unicode
    indicator_count_10 = "2"  # no indicators
    subfield_count_11 = "2"  # no subfields
    base_address_12 = "00000"  # placeholder for base address of data
    encoding_level_17 = "3"  # "3" for abbrieviated level
    cat_conventions_18 = "i"  # ISBD punctuation included
    multipart_indic_19 = Blank()  # no multipart
    field_len_20 = "4"  # placeholder for length of field
    start_character_len_21 = "5"  # placeholder for length of starting character
    implementation_len_22 = "0"  # placeholder for implementation length
    undefined_23 = "0"  # placeholder for undefined length
    success = Field(
        tag,
        i1,
        i2,
        [
            Subfield(
                record_len_00 + record_status_05 + record_type_06 + biblio_level_07
            ),
            type_of_ctrl_08,
            Subfield(
                char_coding_09
                + indicator_count_10
                + subfield_count_11
                + base_address_12
                + encoding_level_17
                + cat_conventions_18
            ),
            multipart_indic_19,
            Subfield(
                field_len_20
                + start_character_len_21
                + implementation_len_22
                + undefined_23
            ),
        ],
    )
    return Result(success, None)


def build_005(record: Record) -> Result:
    """date & time of transaction
    "The date requires 8 numeric characters in the pattern yyyymmdd. The time requires 8 numeric characters in the pattern hhmmss.f, expressed in terms of the 24-hour (00-23) clock."
    """
    tag = 5
    i1, i2 = variable_control_field()
    standard_time = record.timestamp.now(timezone.utc)
    ## NB: python produces this format: YYYY-MM-DD HH:MM:SS.ffffff, e.g. 2020-09-30 12:37:55.713351
    timestamp = str(standard_time).translate(str.maketrans("", "", " -:"))[:16]
    result = Result(Field(tag, i1, i2, [Subfield(timestamp)]), None)
    return result


def build_008(record: Record) -> Result:
    """place & year of pub & main language"""
    tag = 8
    i1, i2 = variable_control_field()
    t = record.timestamp
    date_entered_on_file = Subfield(
        str(t.year)[2:] + str(t.month).zfill(2) + str(t.day).zfill(2)
    )
    pub_status: Subfield = Subfield("s")
    date_1: Subfield = Subfield(record.pub_year)
    date_2: Subfield = Subfield(4 * "|")
    region = check_for_detailed_region(record.country_code, record.state, record.place)
    place_of_pub: list[atoms] = [Subfield(region), *fill_with_blanks(region)]
    books_configuration: list[atoms] = [Subfield(14 * "|"), Blank(), Subfield(2 * "|")]
    lang: list[atoms] = [Subfield(record.langs[0]), *fill_with_blanks(record.langs[0])]
    modified_and_cataloging: Subfield = Subfield(2 * "|")
    content = Content(
        [
            date_entered_on_file,
            pub_status,
            date_1,
            date_2,
            *place_of_pub,
            *books_configuration,
            *lang,
            modified_and_cataloging,
        ]
    )
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_033(record: Record) -> Result:
    """sales dates"""
    tag = 33
    i1 = 0 if len(record.sale_dates) == 1 else 1
    i2 = -1
    result = Result(
        Field(tag, i1, i2, [Subfield(date, "a") for date in record.sale_dates]), None
    )
    return result


def build_040(record: Record) -> Result:
    """Cataloguing source (Oxford)"""
    tag = 40
    i1, i2 = -1, -1
    content = Content(
        [
            Subfield("UkOxU", "a"),
            Subfield("eng", "b"),
            Subfield("rda", "e"),
            Subfield("UkOxU", "c"),
        ]
    )
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_245(record: Record) -> Result:
    """Title
    Field 245 ends with a period, even when another mark of punctuation is present, unless the last word in the field is an abbreviation, initial/letter, or data that ends with final punctuation.
    """
    tag = 245
    i1 = 0
    nonfiling = 0
    error = None
    has_chinese_title = bool(record.title.transliteration)
    if has_chinese_title:
        title, subtitle = record.title.transliteration, record.subtitle.transliteration
        linkage = deal_with_chinese_titles(
            record, record.title.original, record.subtitle.original, i1, nonfiling, tag
        )
    else:
        title, subtitle = record.title.original, record.subtitle.original
        nonfiling, title = check_for_nonfiling(title)
        linkage = None
    i2 = nonfiling
    if title:
        content = Content([])
        if linkage:
            content.add(linkage)
        content.add(Subfield(title, "a"))
        if subtitle:
            content.add([Punctuation(" :"), Subfield(subtitle, "b")])
        if content.can_accept_period():
            content.add(Punctuation("."))

        result = Result(Field(tag, i1, i2, content), error)
    else:
        result = Result(None, (tag, ""))
    return result


def deal_with_chinese_titles(
    record: Record,
    title_original: str,
    subtitle_original: str | None,
    i1: int,
    i2: int,
    tag: int,
) -> Subfield:
    chinese_title = Content(Subfield(title_original, "a"))
    if subtitle_original:
        chinese_title.add([Punctuation(" :"), Subfield(subtitle_original, "b")])
    if tag == 245 and chinese_title.can_accept_period():
        chinese_title.add(Punctuation("."))
    sequence_number = seq_num(record.sequence_number)
    linkage = Subfield(f"880-{sequence_number}", 6)
    build_880(record, chinese_title, i1, i2, tag, sequence_number)
    return linkage


def build_264(record: Record) -> Result:
    """publisher & copyright"""
    tag = 264
    i1 = -1
    i2 = 1  ## "Publication: Field contains a statement relating to the publication, release, or issuing of a resource."
    # i2 = 0  ## "Production: Field contains a statement relating to the production of a resource."
    error = None
    copyright_symbol = "\u00a9"
    place_name = record.place if record.place else record.state
    place = Subfield(place_name, "a")
    publisher = Subfield(record.publisher, "b")
    pub_year = Subfield(
        f"[{record.pub_year}?]" if record.pub_year_is_approx else record.pub_year, "c"
    )
    content = Field(
        tag,
        i1,
        i2,
        [place, Punctuation(" :"), publisher, Punctuation(","), pub_year],
        1,
    )
    if record.copyright:
        # _copyright = Field(tag, i1, 4, [Subfield(f"\u00a9 {record.copyright}", "c")], 2)
        _copyright = Field(
            tag, i1, 4, [Subfield(f"{copyright_symbol} {record.copyright}", "c")], 2
        )
        result = Result([content, _copyright], error)
    else:
        result = Result(content, error)
    return result


def build_300(record: Record) -> Result:
    """physical description"""
    tag = 300
    i1, i2 = -1, -1  ## "undefined"
    pages = Subfield(
        (
            f"approximately {record.extent} pages"
            if record.extent_is_approx
            else f"{record.extent} pages"
        ),
        "a",
    )
    size = Subfield(f"{record.size} cm", "c")
    content = Content([pages, Punctuation(" ;"), size])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_336(record: Record) -> Result:
    """content type (boilerplate)"""
    tag = 336
    content = Content([Subfield("text", "a"), Subfield("rdacontent", 2)])
    result = Result(Field(tag, -1, -1, content), None)
    return result


def build_337(record: Record) -> Result:
    """media type (boilerplate)"""
    tag = 337
    i1, i2 = -1, -1
    content = Content([Subfield("unmediated", "a"), Subfield("rdamedia", 2)])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_338(record: Record) -> Result:
    """carrier type (boilerplate)"""
    tag = 338
    i1, i2 = -1, -1
    content = Content([Subfield("volume", "a"), Subfield("rdacarrier", 2)])
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_876(record: Record) -> Result:
    """
    notes / donations / barcode
    mandatory because of barcode
    """
    tag = 876
    i1, i2 = -1, -1
    content = Content([Subfield(record.barcode, "p")])
    if donation := record.donation:
        content.add(Subfield(donation, "z"))
    if notes := record.hol_notes:
        content.add(Subfield(notes, "z"))
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_904(record: Record) -> Result:
    """authority (boilerplate)"""
    tag = 904
    i1, i2 = -1, -1
    content = Content(Subfield("Oxford Local Record", "a"))
    result = Result(Field(tag, i1, i2, content), None)
    return result


def build_020(record: Record) -> Result:  ##optional
    """isbn (if exists)"""
    tag = 20
    i1, i2 = -1, -1
    contents: list[atoms] = []
    if record.isbn:
        contents.append(Subfield(record.isbn, "a"))
    if record.volume:
        contents.append(Subfield(f"volume {record.volume}", "q"))
    if contents:
        result = Result(Field(tag, i1, i2, Content(contents)), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_024(record: Record) -> Result:  ##optional
    """sales code (if exists)"""
    tag = 24
    i1 = 8
    i2 = -1
    if record.sales_code:
        result = Result(
            Field(tag, i1, i2, Content(Subfield(record.sales_code, "a"))), None
        )
    else:
        result = Result(None, (tag, ""))
    return result


def build_041(record: Record) -> Result:  ##optional
    """language codes if not monolingual"""
    tag = 41
    # i1 = -1  ## "No information...as to whether the item is or includes a translation."
    i1 = 0  ## "Item not a translation/does not include a translation."
    i2 = -1  ## "(followed by) MARC language code"
    is_multi_lingual = len(record.langs) > 1
    if is_multi_lingual:
        ## OPTION 1
        # if record.title.transliteration:
        #     main_lang = record.langs[0]
        #     others = record.langs[1:]
        #     lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # else:
        #     lang_list = f"$a{"$a".join(record.lang)}"
        ## OPTION 2
        # main_lang = record.langs[0]
        # others = record.langs[1:]
        # lang_list = f"$a{"$a".join(others)}$h{main_lang}"
        # result = (build_field(41, [[i1,i2, lang_list]]))
        ## OPTION 3
        result = Result(
            Field(tag, i1, i2, [Subfield(language, "a") for language in record.langs]),
            None,
        )
    else:
        result = Result(None, (tag, ""))
    return result


def build_246(record: Record) -> Result:  ##optional
    """
    Varying Form of Title
    Holds parallel Western title AND/OR original Chinese character title
    NB: Initial articles (e.g., The, La) are generally not recorded in field 246 unless the intent is to file on the article. [https://www.loc.gov/marc/bibliographic/bd246.html]
    """
    tag = 246
    i1 = 3  ## "No note, added entry"
    i2 = 1  ## parallel title
    has_parallel_title = bool(record.parallel_title.original)
    has_chinese_parallel_title = bool(record.parallel_title.transliteration)
    linkage: Subfield | None = None
    if has_chinese_parallel_title:
        parallel_title = record.parallel_title.transliteration
        parallel_subtitle = record.parallel_subtitle.transliteration
        linkage = deal_with_chinese_titles(
            record,
            record.parallel_title.original,
            record.parallel_subtitle.original,
            i1,
            i2,
            tag,
        )
    elif has_parallel_title:  ## (i.e. Western script)
        parallel_title, parallel_subtitle = (
            record.parallel_title.original,
            record.parallel_subtitle.original,
        )
    else:
        parallel_title, parallel_subtitle = "", ""
    if parallel_title:
        parallel_title_sub = Subfield(parallel_title, "a")
        if linkage:
            content = Content([linkage, parallel_title_sub])
        else:
            content = Content(parallel_title_sub)
        if parallel_subtitle:
            content.add(Subfield(parallel_subtitle, "b"))
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_490(record: Record) -> Result:  ## optional
    """Series Statement"""
    tag = 490
    i1 = 0  ## Series not traced: No series added entry is desired for the series.
    i2 = -1
    series_title = Subfield(record.series_title if record.series_title else "", "a")
    series_enum = Subfield(record.series_enum if record.series_enum else "", "v")
    content: Content | None = None
    if record.series_title and record.series_enum:
        content = Content([series_title, Punctuation(" ;"), series_enum])
    elif record.series_title:
        content = Content(series_title)
    elif record.series_enum:
        content = Content(series_enum)
    if content:
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


def build_500(record: Record) -> Result:  ##optional
    """general notes
    Punctuation - Field 500 ends with a period unless another mark of punctuation is present. If the final subfield is subfield $5, the mark of punctuation precedes that subfield.
    """
    tag = 500
    i1 = -1
    i2 = -1
    if record.notes:
        content = Content(Subfield(record.notes, "a"))
        if content.can_accept_period():
            content.add(Punctuation("."))
        result = Result(Field(tag, i1, i2, content), None)
    else:
        result = Result(None, (tag, ""))
    return result


# TODO: need an item with both a Chinese title and subtitle to test the sequence number logic.
def build_880(
    record: Record, title: Content, i1: int, i2: int, caller: int, sequence_number: str
) -> None:  ##optional
    """Alternate Graphic Representation
    NB. unlike the other fields, this isn't called directly but by the linked field
    looks like: =880  31$6246-01$a中國書畫、陶瓷及藝術品拍賣會
    """
    record.sequence_number += 1
    content = Content(Subfield(f"{str(caller).zfill(3)}-{sequence_number}", 6))
    content.add(title.contents)
    line = Field(880, i1, i2, content)
    record.links.append(line)


def check_if_mandatory(result: Result, is_mandatory: bool) -> list[Field] | None:
    """
    silently suppresses optional fields if empty;
    stops with error if required field is empty
    """
    output: list[Field] | None = None
    if result.is_err:
        numeric_tag, error_msg = result.is_err
        if is_mandatory:
            msg = (
                error_msg
                if error_msg
                else f"Data for required field {str(numeric_tag).zfill(3)} is required."
            )
            logger.warning(msg)
            raise MissingFieldError(msg)
    elif result.is_ok:
        if isinstance(result.is_ok, list):
            output = result.is_ok
        else:
            output = [result.is_ok]
    return output


def line_prefix(numeric_tag: int) -> str:
    display_tag = "LDR" if numeric_tag == 0 else seq_num(numeric_tag)
    return f"={expand_tag(display_tag)}  "


def expand_tag(field_number: str) -> str:
    return str(field_number).zfill(3)


def seq_num(sequence_number: int) -> str:
    return str(sequence_number).zfill(2)


def check_for_nonfiling(title: str, lang: str = "eng") -> tuple[int, str]:
    """
    Check for manual nonfiling indicator (@@) & returns its position + extracts it from title string
    if no manual indication, check for nonfiling words according to language of title"""
    nonfiling_words = {
        "eng": ("the", "a", "an"),
        "fr": ("le", "la", "les", "l'", "un", "une"),
        "it": ("lo", "il", "i", "l'", "gli", "le", "un", "una", "un'"),
        "sp": ("el", "la", "las", "los", "un", "una", "unos", "unas"),
        "gw": ("der", "die", "das", "ein", "eine"),
        "ne": ("de", "het"),
        "sw": ("en", "ett", "den", "det", "de"),
        "dk": ("en", "et"),
        "no": ("en", "ei", "et"),
        "po": ("o", "a", "os", "as", "um", "uma", "uns", "umas"),
        "chi": (),
    }
    break_char = "@@"
    nonfiling = title.find(break_char)
    result = (0, title)
    if nonfiling > 0:
        result = (nonfiling, title.replace(break_char, "", 1))
    else:
        for article in nonfiling_words[lang]:
            test = re.match(f"({article}\\s?[^\\w\\s]?)\\w", title, re.I)
            if test:
                result = (test.span()[1] - 1, title)
                break
    return result


def variable_control_field():
    return (-2, -2)


def build_marc_records(records: list[Record]) -> list[marc_record]:
    # marc_records: list[list[Field]] = []
    marc_records: list[marc_record] = []
    for record in records:
        # print(record, type(record))
        # pprint(record)
        marc = apply_marc_logic(record)
        marc_records.append(marc)
    return marc_records


def apply_marc_logic(record: Record) -> list[Field]:
    marc_record: list[Field] = []
    # print(record, type(record))
    # fields_to_deploy: tuple[tuple[Callable[[Record], Result], bool]] = (
    fields_to_deploy: tuple[tuple[Callable, bool], ...] = (
        (build_leader, True),
        (build_040, True),
        (build_336, True),
        (build_337, True),
        (build_338, True),
        (build_904, True),
        (build_005, True),
        (build_008, True),
        (build_033, True),
        (build_245, True),
        (build_264, True),
        (build_300, True),
        (build_490, False),  # series statement
        (build_876, True),
        (build_020, False),  # isbn
        (build_024, False),  # sales code
        (build_041, False),  # language if not monolingual
        (build_246, False),  # parallel title
        (build_500, False),  # general notes
    )
    for builder, is_mandatory in fields_to_deploy:
        field = check_if_mandatory(builder(record), is_mandatory)
        if field:
            for repeat in field:
                marc_record.append(repeat)
    if len(record.links):
        for link in record.links:
            if link:
                marc_record.append(link)
    marc_record.sort(key=lambda x: (x.tag * 10) + x.ordering)
    return marc_record


def write_mrk_files(data: list[list[str]], file_name: str = "output.mrk") -> None:
    out_file = output_file_dir / file_name
    with open(out_file, "w", encoding="utf-8") as f:
        # print(f">>>> entries: {len(data)}")
        last_record: int = len(data) - 1
        for i, record in enumerate(data):
            for field in record:
                f.write(field + "\n")
            if i < last_record:
                f.write("\n")


def write_mrc_binaries(data: list[list[Field]], file_name: str = "output.mrc") -> None:
    out_file = output_file_dir / file_name
    flat_records = make_binary(data)
    with open(out_file, "wb") as f:
        for line in flat_records:
            f.write(line)


def make_binary(data: list[list[Field]]) -> list[bytes]:
    # GS = "@"
    # RS = "%"
    # US = "#"
    GS = chr(29)
    RS = chr(30)
    # output: list[str] = []
    output: list[bytes] = []
    tmp = []
    for record in data:
        leader: Field = record[0]
        fields = []
        current_start_position = 0
        for field in record[1:]:
            tag = field.tag
            # contents = field.to_mrc()
            contents = field.serialize("mrc")
            line_length_with_final_record_separator = len(contents.encode("utf-8"))
            fields.append(
                (
                    tag,
                    contents,
                    line_length_with_final_record_separator,
                    current_start_position,
                )
            )
            current_start_position += line_length_with_final_record_separator
        # tmp.append((leader.to_mrc(), fields))
        tmp.append((leader.serialize("mrc"), fields))

        directory = [(f[0], f[2], f[3]) for f in fields]
        # leader_length = int(24)
        leader_length: int = 24
        directory_length: int = (12 * len(directory)) + 1
        base_address_of_data: int = leader_length + directory_length
        logical_record_length: int = base_address_of_data + directory[-1][2]
        leader_mrc = f"{str(logical_record_length).zfill(5)}nam a22{str(base_address_of_data).zfill(5)}3i 4500"
        directory_mrc = (
            "".join(
                [
                    f"{str(f[0]).zfill(3)}{str(f[1]).zfill(4)}{str(f[2]).zfill(5)}"
                    for f in directory
                ]
            )
            + RS
        )
        fields_mrc = "".join([f[1] for f in fields])
        binary_line = f"{leader_mrc}{directory_mrc}{fields_mrc}{GS}".encode("utf-8")
        output.append(binary_line)
    return output


def parse_file_into_rows(
    file_path: Path,
) -> tuple[list[str], list[worksheet_row]]:
    is_excel_file = file_path.suffix.startswith(".xl")
    if is_excel_file:
        excel_file_name = str(file_path.resolve())
        worksheet = load_workbook(filename=excel_file_name).active
        headers, raw_rows = extract_from_excel(worksheet)
    else:
        headers, raw_rows = extract_from_csv(file_path)
    return (headers, raw_rows)


def extract_from_csv(file_address: Path) -> tuple[list[str], list[worksheet_row]]:
    sheet = []
    headers = []
    delimiter = "," if file_address.suffix == ".csv" else "\t"
    with open(file_address.resolve(), mode="r", encoding="utf-8") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        for i, row in enumerate(csv_reader):
            row = normalize_row(row)
            if i == 0 and settings.first_row_is_header:
                headers = row
            else:
                sheet.append(row)
    return (headers, sheet)


def flatten_fields_to_strings(input: list[marc_record]) -> list[list[str]]:
    output: list[list[str]] = []
    for i, record in enumerate(input):
        # output.append([field.to_string() for field in record])
        output.append([field.serialize() for field in record])
    return output


def write_marc_files(records: list[marc_record], excel_file_address: Path) -> None:
    records_with_string_fields = flatten_fields_to_strings(records)
    print(f"Writing {len(records)} record(s)...")
    write_mrk_files(records_with_string_fields, f"{excel_file_address.stem}.mrk")
    write_mrc_binaries(records, f"{excel_file_address.stem}.mrc")


def run() -> None:
    # for file in Path(settings.data_dir).glob("*.xls[xm]"):
    extensions = ["*.xlsx", "*.xlsm", "*.csv", "*.tsv"]
    file_list: list[Path] = []
    for ext in extensions:
        file_list.extend(Path(settings.data_dir).glob(ext))

    print(f"There {len(file_list)} files to process.")
    for file in file_list:
        # is_excel_file = file.suffix.startswith(".xl")
        logger.info(f"\n>>>>> processing: {file.name}")
        print(
            f">>>>> now processing: {file}>{file.name} ({file.suffix})"
        )
        headers, raw_rows = parse_file_into_rows(file)
        # if is_excel_file:
        #     headers, raw_rows = parse_file_into_rows(file)
        # else:
        #     headers, raw_rows = extract_from_csv(file)
        records = parse_rows_into_records(raw_rows)
        del headers
        del raw_rows
        marc_records = build_marc_records(records)
        del records
        write_marc_files(marc_records, file)


def main() -> None:
    run()


if __name__ == "__main__":
    # logger = logging.getLogger(__name__)
    main()
